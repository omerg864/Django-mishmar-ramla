import datetime
from datetime import date as Date
from datetime import time as Time
import io
import random
from time import time
from click import pass_obj
import xlsxwriter as xlsxwriter
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.models import User, Group
from django.http import FileResponse
from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect
from django.template.defaulttags import register
from django.views.generic import UpdateView, ListView, DetailView, CreateView
from .backend.Schedule.Organizer import Organizer
from .forms import SettingsForm, ShiftForm, ShiftViewForm, WeekUpdateForm, ShiftWeekForm, ShiftWeekViewForm
from django.forms.models import model_to_dict
from .models import Post, ValidationLog
from .models import Settings3 as Settings
from .models import Shift1 as Shift
from .models import Event
from .models import Organization as Organization
from .models import Week
from .models import ShiftWeek
from .models import Arming_Log
from .models import Gun
from users.models import UserSettings as USettings
import openpyxl
from django.utils import timezone
from openpyxl.utils import get_column_letter
import requests
from deep_translator import GoogleTranslator
import os
from django.views.generic.dates import DayArchiveView, MonthArchiveView


EMPTY_SIGNATURE = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAASwAAACgCAYAAAC2eFFiAAAAAXNSR0IArs4c6QAABKtJREFUeF7t1AEJAAAMAsHZv/RyPNwSyDncOQIECEQEFskpJgECBM5geQICBDICBitTlaAECBgsP0CAQEbAYGWqEpQAAYPlBwgQyAgYrExVghIgYLD8AAECGQGDlalKUAIEDJYfIEAgI2CwMlUJSoCAwfIDBAhkBAxWpipBCRAwWH6AAIGMgMHKVCUoAQIGyw8QIJARMFiZqgQlQMBg+QECBDICBitTlaAECBgsP0CAQEbAYGWqEpQAAYPlBwgQyAgYrExVghIgYLD8AAECGQGDlalKUAIEDJYfIEAgI2CwMlUJSoCAwfIDBAhkBAxWpipBCRAwWH6AAIGMgMHKVCUoAQIGyw8QIJARMFiZqgQlQMBg+QECBDICBitTlaAECBgsP0CAQEbAYGWqEpQAAYPlBwgQyAgYrExVghIgYLD8AAECGQGDlalKUAIEDJYfIEAgI2CwMlUJSoCAwfIDBAhkBAxWpipBCRAwWH6AAIGMgMHKVCUoAQIGyw8QIJARMFiZqgQlQMBg+QECBDICBitTlaAECBgsP0CAQEbAYGWqEpQAAYPlBwgQyAgYrExVghIgYLD8AAECGQGDlalKUAIEDJYfIEAgI2CwMlUJSoCAwfIDBAhkBAxWpipBCRAwWH6AAIGMgMHKVCUoAQIGyw8QIJARMFiZqgQlQMBg+QECBDICBitTlaAECBgsP0CAQEbAYGWqEpQAAYPlBwgQyAgYrExVghIgYLD8AAECGQGDlalKUAIEDJYfIEAgI2CwMlUJSoCAwfIDBAhkBAxWpipBCRAwWH6AAIGMgMHKVCUoAQIGyw8QIJARMFiZqgQlQMBg+QECBDICBitTlaAECBgsP0CAQEbAYGWqEpQAAYPlBwgQyAgYrExVghIgYLD8AAECGQGDlalKUAIEDJYfIEAgI2CwMlUJSoCAwfIDBAhkBAxWpipBCRAwWH6AAIGMgMHKVCUoAQIGyw8QIJARMFiZqgQlQMBg+QECBDICBitTlaAECBgsP0CAQEbAYGWqEpQAAYPlBwgQyAgYrExVghIgYLD8AAECGQGDlalKUAIEDJYfIEAgI2CwMlUJSoCAwfIDBAhkBAxWpipBCRAwWH6AAIGMgMHKVCUoAQIGyw8QIJARMFiZqgQlQMBg+QECBDICBitTlaAECBgsP0CAQEbAYGWqEpQAAYPlBwgQyAgYrExVghIgYLD8AAECGQGDlalKUAIEDJYfIEAgI2CwMlUJSoCAwfIDBAhkBAxWpipBCRAwWH6AAIGMgMHKVCUoAQIGyw8QIJARMFiZqgQlQMBg+QECBDICBitTlaAECBgsP0CAQEbAYGWqEpQAAYPlBwgQyAgYrExVghIgYLD8AAECGQGDlalKUAIEDJYfIEAgI2CwMlUJSoCAwfIDBAhkBAxWpipBCRAwWH6AAIGMgMHKVCUoAQIGyw8QIJARMFiZqgQlQMBg+QECBDICBitTlaAECBgsP0CAQEbAYGWqEpQAAYPlBwgQyAgYrExVghIgYLD8AAECGQGDlalKUAIEDJYfIEAgI2CwMlUJSoCAwfIDBAhkBAxWpipBCRAwWH6AAIGMgMHKVCUoAQIPBfEAoeZvv+sAAAAASUVORK5CYII="


default_language = os.environ.get("DEFAULT_LANGUAGE")

if len(Settings.objects.all()) == 0:
    new_settings = Settings(submitting=True, pin_code=1234, officer="", city="", max_seq0=2, max_seq1=2)
    new_settings.save()

base_strings = {1: "הגשת משמרות", 2: "סידור", 3: "סידורים", 4: "הגדרות", 5: "ניהול נתונים", 6: "פרופיל", 7: "התנתק",
                8: "התחבר", 9: "הירשם"}


@staff_member_required
def settings_view(request):
    settings = Settings.objects.all().last()
    checked = settings.submitting
    if request.method == 'POST':
        checked = request.POST.get("serv")
        if checked:
            checked = False
        else:
            checked = True
        success, fail = ['שינויים נשמרו!', 'שינויים לא נשמרו!']
        settings_form = SettingsForm(request.POST, instance=settings)
        settings_form.instance.submitting = checked
        if settings_form.is_valid():
            messages.success(request, translate_text(success, request.user, "hebrew"))
            settings_form.save()
        else:
            messages.error(request, translate_text(fail, request.user, "hebrew"))
    else:
        settings_form = SettingsForm(instance=settings)
    context = {
        "settings_form": settings_form,
        "checked": checked,
        "base": base_strings
    }
    return render(request, "Schedule/settings.html", context)


def home(request):
    settings = Settings.objects.all().first()
    data = {}
    api_key = "4cba4792d5c0c0222cc84e409138af7a"
    base_url = "http://api.openweathermap.org/data/2.5/weather?"
    if settings.city == '':
        city_name = "Ramla"
    else:
        translator2en = GoogleTranslator(source='auto', target='en')
        city_name = translator2en.translate(settings.city)
    try:
        complete_url = base_url + "appid=" + api_key + "&q=" + city_name
        response = requests.get(complete_url)
        data = response.json()
    except:
        data = {"Not Found:": ""}

    if data["cod"] != "404":
        try:
            y = data["main"]
            current_temperature = str(int(y["temp"] - 273.15)) + " °C"
            current_pressure = str(y["pressure"]) + " hPa"
            current_humidiy = str(y["humidity"]) + "%"
            weather_description = data["weather"][0]["description"]
            weather = {
                translate_text("טמפרטורה", request.user, "hebrew"): current_temperature,
                translate_text("לחץ אטמוספרי", request.user, "hebrew"): current_pressure,
                translate_text("לחות", request.user, "hebrew"): current_humidiy,
                translate_text("תיאור", request.user, "hebrew"):
                    translate_text(weather_description, request.user, "english")
            }
        except AttributeError:
            print("Weather Error")
            weather = {
                translate_text("לא נמצא", request.user, "hebrew"):
                    translate_text("לא ניתן לטעון מזג האוויר", request.user, "hebrew")
            }
    else:
        print(" City Not Found ")
        weather = {
            translate_text("לא נמצא", request.user, "hebrew"): translate_text("עיר לא נמצא", request.user, "hebrew")
        }
    posts = Post.objects.all()
    context = {
        "weather": weather,
        "posts": posts,
        "city": city_name,
    }
    return render(request, "Schedule/Home.html", context)


def error_404_view(request, exception):
    return render(request, 'Schedule/404.html')


class ArmingDayView(LoginRequiredMixin, DayArchiveView):
    queryset = Arming_Log.objects.all()
    date_field = "date"
    allow_future = True
    allow_empty = True
    template_name = "Schedule/arming.html"
    ordering = ["shift_num", "time_in"]

    def get_context_data(self, **kwargs):
        ctx = super(ArmingDayView, self).get_context_data(**kwargs)
        guns = Gun.objects.all()
        user_name = self.request.user.first_name + " " + self.request.user.last_name
        months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
        date1 = Date(self.kwargs['year'], months.index(getmonth(self.kwargs['month'].lower())) + 1, self.kwargs['day'])
        validation_log = ValidationLog.objects.all().filter(date=date1).first()
        num_mags_list = [1, 2, 3]
        hand_cuffs_list = [6, 1, 2, 3, 4, 5, 7, 8]
        mag_case_list = [6, 1, 2, 3, 4, 5, 7]
        gun_case_list = [6, 1, 2, 3, 4, 5, 7, 8]
        ctx["num_mags_list"] = num_mags_list
        ctx["hand_cuffs_list"] = hand_cuffs_list
        ctx["mag_case_list"] = mag_case_list
        ctx["gun_case_list"] = gun_case_list
        ctx["guns"] = guns
        ctx["user_name"] = user_name
        ctx["validation_log"] = validation_log
        return ctx
    
    def post(self, request, *args, **kwargs):
        shift = 0
        if "add" in request.POST:
            request.session["gun_id"] = request.POST.get(f"guns")
            if request.user.username != "metagber": 
                request.session["name"] = request.user.first_name + " " + request.user.last_name
            else:
                request.session["name"] = request.POST.get(f"user_name")
            session_keyes = ["id_num", "time_in", "num_mags", "hand_cuffs", "gun_case", "mag_case", "keys", "radio", "radio_kit"]
            int_keyes = ["num_mags", "hand_cuffs", "gun_case", "mag_case"]
            bool_keyes = ["keys", "radio", "radio_kit"]
            for key in session_keyes:
                if key in int_keyes:
                    request.session[key] = int(request.POST[f"{key}"])
                elif key in bool_keyes:
                    request.session[key] = checkbox(request.POST.get(f"{key}", None))
                else:
                    request.session[key] = request.POST[f"{key}"]
            request.session["shift_num"] = int(request.POST.get("shifts"))
            time_out = request.POST.get("time_out")
            if request.user.groups.filter(name="manager").exists():
                request.session["reqtype"] = "manager"
            else:
                request.session["reqtype"] = "add"
            if time_out != "":
                request.session["time_out"] = time_out
            else:
                request.session["time_out"] = ""
            months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
            request.session["year"] = self.kwargs['year']
            request.session["month"] = months.index(getmonth(self.kwargs['month'].lower())) + 1
            request.session["day"] = self.kwargs['day']
            messages.success(request, "הנתונים הועברו בהצלחה")
            return redirect("arming-new")
        elif "change" in request.POST:
            log = Arming_Log.objects.get(id=request.POST.get("change"))
            request.session["gun_id"] = request.POST.get(f"guns{log.id}")
            request.session["shift_num"] = int(request.POST.get(f"shifts{log.id}"))
            session_keyes = ["id_num", "time_in", "num_mags", "hand_cuffs", "gun_case", "mag_case", "keys", "radio", "radio_kit"]
            int_keyes = ["num_mags", "hand_cuffs", "gun_case", "mag_case"]
            bool_keyes = ["keys", "radio", "radio_kit"]
            for key in session_keyes:
                if key in int_keyes:
                    request.session[key] = int(request.POST[f"{key}{log.id}"])
                elif key in bool_keyes:
                    request.session[key] = checkbox(request.POST.get(f"{key}{log.id}", None))
                else:
                    request.session[key] = request.POST[f"{key}{log.id}"]
            time_out = request.POST.get(f"time_out{log.id}")
            request.session["log_id"] = log.id
            if request.user.username != "metagber":
                name = request.user.first_name + " " + request.user.last_name
            else:
                name = request.POST.get(f"user_name{log.id}")
            request.session["name"] = log.name
            if name == log.name and request.user.groups.filter(name="manager").exists():
                request.session["reqtype"] = "change manager"
            elif name == log.name:
                request.session["reqtype"] = "change"
            else:
                request.session["reqtype"] = "validation"
            if time_out != "":
                request.session["time_out"] = time_out
            else:
                request.session["time_out"] = ""
            messages.info(request, " הנתונים הועברו בהצלחה כדי לשמור יש לחתום")
            return redirect("signature", log.id)
        elif "month_log" in request.POST:
            return redirect("armingmonth", year=self.kwargs['year'], month=self.kwargs['month'])
        elif "shift1" in request.POST:
            shift = 1
        elif "shift2" in request.POST:
            shift = 2
        elif "shift3" in request.POST:
            shift = 3
        elif "goto" in request.POST:
            date = request.POST.get("goto_date")
            print(date)
            print(type(date))
            date1 = Date(int(date[0:4]), int(date[5:7]), int(date[8:10]))
            months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
            return redirect('armingday', year=date1.year, month=getmonth(date1.strftime("%b")), day=date1.day)
        if shift != 0:
            manager = request.POST.get(f"manager{shift}")
            if manager == "":
                messages.info(request, translate_text("נא למלא את שם המנהל", request.user, "hebrew"))
                return HttpResponseRedirect(request.path_info)
            val_logs = ValidationLog.objects.all()
            months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
            request.session["year"] = self.kwargs['year']
            request.session["month"] = months.index(getmonth(self.kwargs['month'].lower())) + 1
            request.session["day"] = self.kwargs['day']
            date1 = Date(self.kwargs['year'], months.index(getmonth(self.kwargs['month'].lower())) + 1, self.kwargs['day'])
            log = val_logs.filter(date=date1)
            if len(log) == 0:
                request.session["reqtype"] = "add"
            else:
                request.session["reqtype"] = "change"
            session_keyes = ["gun_safe", "gun_shift", "time", "manager"]
            for key in session_keyes:
                request.session[key] = request.POST[f"{key}{shift}"]
            request.session["shift"] = shift
            messages.info(request, " הנתונים הועברו בהצלחה כדי לשמור יש לחתום")
            return redirect("validation-signature")

class ArmingLogUpdate(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Arming_Log
    template_name = "Schedule/signature_page.html"
    fields = ('valid_in', 'valid_out', 'signature_in', 'signature_out')
    context_object_name = 'arming'

    def test_func(self):
        name = self.request.user.first_name + " " + self.request.user.last_name
        if self.get_object().name == name or self.request.user.groups.filter(name="manager").exists() or self.request.user.username == "metagber":
            return True
        return False

    def get_context_data(self, **kwargs):
        ctx = super(ArmingLogUpdate, self).get_context_data(**kwargs)
        guns = Gun.objects.all()
        user_name = self.request.session["name"]
        num_mags_list = [1, 2, 3]
        hand_cuffs_list = [6, 1, 2, 3, 4, 5, 7, 8]
        mag_case_list = [6, 1, 2, 3, 4, 5, 7]
        gun_case_list = [6, 1, 2, 3, 4, 5, 7, 8]
        ctx["num_mags_list"] = num_mags_list
        ctx["hand_cuffs_list"] = hand_cuffs_list
        ctx["mag_case_list"] = mag_case_list
        ctx["gun_case_list"] = gun_case_list
        ctx["guns"] = guns
        ctx["user_name"] = user_name
        session_keyes = ["gun_id","name", "shift_num", "id_num", "time_in", "num_mags", "hand_cuffs", "gun_case", "mag_case", "keys", "radio", "radio_kit", "time_out", "reqtype"]
        for key in session_keyes:
            ctx[key] = self.request.session[key]
        gun_id = self.request.session["gun_id"]
        short_name = guns.filter(id=gun_id).first().short_name
        ctx["short_name"] = short_name
        return ctx
    
    def post(self, request, *args, **kwargs):
        reqtype = request.session["reqtype"]
        session_keyes = ["gun_id", "shift_num", "id_num", "time_in", "num_mags", "hand_cuffs", "gun_case", "mag_case", "keys", "radio", "radio_kit", "time_out", "reqtype"]
        log = Arming_Log.objects.filter(id=self.get_object().id).first()
        gun_id = request.session["gun_id"]
        gun  = Gun.objects.filter(id=gun_id).first()
        log.gun = gun
        log.name = request.session["name"]
        log.shift_num = request.session["shift_num"]
        log.id_num = request.session["id_num"]
        log.time_in = request.session["time_in"]
        log.num_mags = request.session["num_mags"]
        log.hand_cuffs = request.session["hand_cuffs"]
        log.gun_case = request.session["gun_case"]
        log.mag_case = request.session["mag_case"]
        log.keys = request.session["keys"]
        log.radio = request.session["radio"]
        log.radio_kit = request.session["radio_kit"]
        if request.session["time_out"] != "":
            log.time_out = request.session["time_out"]
        sig_in = request.POST.get('sig-dataUrl')
        sig_out = request.POST.get('sig-dataUrl_out')
        print(sig_in)
        print(sig_out)
        if reqtype == "change manager":
            valid_in = request.POST.get('sig-dataUrl_valid')
            valid_out = request.POST.get('sig-dataUrl_out_valid')
            if sig_in != EMPTY_SIGNATURE:
                log.signature_in = sig_in
            if sig_out != EMPTY_SIGNATURE:
                log.signature_out = sig_out
            if valid_in != EMPTY_SIGNATURE:
                log.valid_in = valid_in
            if valid_out != EMPTY_SIGNATURE:
                log.valid_out = valid_out
        elif reqtype == "change":
            if sig_in != EMPTY_SIGNATURE:
                log.signature_in = sig_in
            if sig_out != EMPTY_SIGNATURE:
                log.signature_out = sig_out
        else:
            if sig_in != EMPTY_SIGNATURE:
                log.valid_in = sig_in
            if sig_out != EMPTY_SIGNATURE:
                log.valid_out = sig_out
        if sig_in == EMPTY_SIGNATURE and sig_out == EMPTY_SIGNATURE and reqtype != "change manager":
            messages.warning(request, "אנא הכנס את החתימה שלך")
            return HttpResponseRedirect(request.path_info)
        elif reqtype == "change manager" and sig_in == EMPTY_SIGNATURE and sig_out == EMPTY_SIGNATURE and valid_in == EMPTY_SIGNATURE and valid_out == EMPTY_SIGNATURE:
            messages.warning(request, "אנא הכנס את החתימה שלך")
            return HttpResponseRedirect(request.path_info)
        else:
            for key in session_keyes:
                del request.session[key]
            log.save()
            messages.success(request, "הנתונים נשמרו בהצלחה")
            return redirect('armingday', year=int(self.get_object().date.strftime("%Y")), month=self.get_object().date.strftime("%b"), day=int(self.get_object().date.strftime("%d")))
        
class ArmingCreateView(LoginRequiredMixin, CreateView):
    model = Arming_Log
    template_name = "Schedule/signature_create.html"
    fields = "__all__"

    def get_context_data(self, **kwargs):
        ctx = super(ArmingCreateView, self).get_context_data(**kwargs)
        guns = Gun.objects.all()
        user_name = self.request.user.first_name + " " + self.request.user.last_name
        num_mags_list = [1, 2, 3]
        hand_cuffs_list = [6, 1, 2, 3, 4, 5, 7, 8]
        mag_case_list = [6, 1, 2, 3, 4, 5, 7]
        gun_case_list = [6, 1, 2, 3, 4, 5, 7, 8]
        ctx["num_mags_list"] = num_mags_list
        ctx["hand_cuffs_list"] = hand_cuffs_list
        ctx["mag_case_list"] = mag_case_list
        ctx["gun_case_list"] = gun_case_list
        ctx["guns"] = guns
        ctx["user_name"] = user_name
        session_keyes = ["gun_id", "name", "shift_num", "id_num", "time_in", "num_mags", "hand_cuffs", "gun_case", "mag_case", "keys", "radio", "radio_kit", "time_out", "reqtype"]
        for key in session_keyes:
            ctx[key] = self.request.session[key]
        ctx["date"] = Date(self.request.session["year"], self.request.session["month"], self.request.session["day"])
        ctx["gun_s"] = Gun.objects.filter(id=self.request.session["gun_id"]).first()
        return ctx
    
    def post(self, request, *args, **kwargs):
        session_keyes = ["gun_id", "name", "shift_num", "id_num", "time_in", "num_mags", "hand_cuffs", "gun_case", "mag_case", "keys", "radio", "radio_kit", "time_out", "reqtype"]
        reqtype = request.session["reqtype"]
        name = request.session["name"]
        gun_id = request.session["gun_id"]
        gun  = Gun.objects.filter(id=gun_id).first()
        gun = gun
        shift_num = request.session["shift_num"]
        id_num = request.session["id_num"]
        time_in = request.session["time_in"]
        num_mags = request.session["num_mags"]
        hand_cuffs = request.session["hand_cuffs"]
        gun_case = request.session["gun_case"]
        mag_case = request.session["mag_case"]
        keys = request.session["keys"]
        radio = request.session["radio"]
        radio_kit = request.session["radio_kit"]
        time_out = request.session["time_out"]
        date1 = Date(request.session["year"], request.session["month"], request.session["day"])
        new_log  = Arming_Log(name=name, id_num=id_num, shift_num=shift_num, date=date1, time_in=time_in,
        gun=gun, num_mags=num_mags, hand_cuffs=hand_cuffs, gun_case=gun_case, mag_case=mag_case, keys=keys,
        radio=radio, radio_kit=radio_kit)
        if time_out != "":
            new_log.time_out = time_out
        sig_in = request.POST.get('sig-dataUrl')
        sig_out = request.POST.get('sig-dataUrl_out')
        signature = False
        if sig_in != EMPTY_SIGNATURE:
            new_log.signature_in = sig_in
            signature = True
        if sig_out != EMPTY_SIGNATURE:
            new_log.signature_out = sig_out
            signature = True
        if reqtype == "manager":
            valid_in = request.POST.get('sig-dataUrl_valid')
            valid_out = request.POST.get('sig-dataUrl_out_valid')
            if valid_in != EMPTY_SIGNATURE:
                new_log.valid_in = request.POST.get('sig-dataUrl_valid')
            if valid_out != EMPTY_SIGNATURE:
                new_log.valid_out = request.POST.get('sig-dataUrl_out_valid')
        if signature:
            for key in session_keyes:
                del request.session[key]
            new_log.save()
            return redirect('armingday', year=int(date1.strftime("%Y")), month=date1.strftime("%b"), day=int(date1.strftime("%d")))
        else:
            messages.warning(request, "אנא הכנס את החתימה שלך")
            return HttpResponseRedirect(request.path_info)

def Validation_Log_Signature(request):
    context = {}
    session_keyes = ["gun_safe", "gun_shift", "time", "manager", "reqtype", "shift", "year", "month", "day"]
    for key in session_keyes:
        context[key] = request.session[key]
    date1 = Date(int(context["year"]), int(context["month"]), int(context["day"]))
    context["date"] = date1
    if request.method == "POST":
        sig = request.POST.get('sig-dataUrl')
        if sig == EMPTY_SIGNATURE:
            messages.warning(request, "אנא הכנס את החתימה שלך")
            return HttpResponseRedirect(request.path_info)
        else:
            shift = context["shift"]
            if shift == 1:
                shift = "m"
            elif shift == 2:
                shift = "a"
            else:
                shift = "n"
            if context["reqtype"] == "add":
                log = ValidationLog()
                log.date = date1
            else:
                log = ValidationLog.objects.filter(date=date1).first()
            log.__setattr__(f"num_guns_safe_{shift}", context["gun_safe"])
            log.__setattr__(f"num_guns_shift_{shift}", context["gun_shift"])
            log.__setattr__(f"time_checked_{shift}", context["time"])
            log.__setattr__(f"name_checked_{shift}", context["manager"])
            log.__setattr__(f"sig_{shift}", sig)
            for key in session_keyes:
                del request.session[key]
            log.save()
            messages.success(request, "החתימה נשמרה בהצלחה")
            return redirect('armingday', year=int(date1.strftime("%Y")), month=date1.strftime("%b"), day=int(date1.strftime("%d")))
    return render(request, "Schedule/validation_signature.html", context)



class ArmingMonthView(LoginRequiredMixin, MonthArchiveView):
    queryset = Arming_Log.objects.all()
    date_field = "date"
    allow_future = True
    allow_empty = True
    template_name = "Schedule/arming-month.html"
    ordering = ["date", "time_in"]

    def get_context_data(self, **kwargs):
        ctx = super(ArmingMonthView, self).get_context_data(**kwargs)
        guns = Gun.objects.all()
        user_name = self.request.user.first_name + " " + self.request.user.last_name
        num_mags_list = [1, 2, 3]
        hand_cuffs_list = [6, 1, 2, 3, 4, 5, 7, 8]
        mag_case_list = [6, 1, 2, 3, 4, 5, 7]
        gun_case_list = [6, 1, 2, 3, 4, 5, 7, 8]
        ctx["num_mags_list"] = num_mags_list
        ctx["hand_cuffs_list"] = hand_cuffs_list
        ctx["mag_case_list"] = mag_case_list
        ctx["gun_case_list"] = gun_case_list
        ctx["guns"] = guns
        ctx["user_name"] = user_name
        return ctx
    
    def post(self, request, *args, **kwargs):
        pass


def checkbox(value):
    print(value)
    if value == "on":
        return True
    return False


@login_required
def shift_view(request):
    organization = Organization.objects.order_by('-date')[0]
    shifts_weeks_served = ShiftWeek.objects.all().filter(date=organization.date)
    forms = []
    for i in range(organization.num_weeks):
        forms.append("")
    form = None
    notes_text = ""
    empty = False
    settings = Settings.objects.last()
    submitting = settings.submitting
    user_settings = USettings.objects.all().filter(user=request.user).first()
    days = []
    for x in range(organization.num_weeks * 7):
        days.append(Organization.objects.order_by('-date')[0].date + datetime.timedelta(days=x))
    events = Event.objects.all()
    for x in range(organization.num_weeks * 7):
        if len(events.filter(date2=days[x])) > 0:
            for ev in events.filter(date2=days[x]):
                if user_settings.nickname == ev.nickname:
                    message = f'לא לשכוח בתאריך {ev.date2} יש {ev.description}. אם יש שינוי להודיע!'
                    message = translate_text(message, request.user, "hebrew")
                    messages.info(request, message)
                elif ev.nickname == 'כולם':
                    message = f'לא לשכוח בתאריך {ev.date2} יש {ev.description}'
                    message = translate_text(message, request.user, "hebrew")
                    messages.info(request, message)
    if request.method == 'POST':
        if not already_submitted(request.user):
            #form = ShiftForm(request.POST)
            shift = Shift()
            for i in range(organization.num_weeks):
                new_form = ShiftWeekForm(request.POST)
                forms[i] = new_form
        else:
            last_date = Organization.objects.order_by('-date')[0].date
            shifts = Shift.objects.filter(date=last_date)
            shift = shifts.filter(username=request.user).first()
            notes_text = str(shift.notes)
            #form = ShiftForm(request.POST, instance=shift)
            weeks = shifts_weeks_served.filter(username=request.user).order_by('num_week')
            for week in weeks:
                new_form = ShiftWeekForm(request.POST, instance=week)
                forms[week.num_week] = new_form
        shift.username = request.user
        shift.date = Organization.objects.order_by('-date')[0].date
        notes_area = request.POST.get("notesArea")
        shift.notes = notes_area
        shift.seq_night = request.POST.get(f"seq_night", 0)
        shift.seq_noon = request.POST.get(f"seq_noon", 0)
        error = False
        shift.save()
        if already_submitted(request.user):
            for form in forms:
                if not form.is_valid():
                    error = True
            if not error:
                for j in range(len(forms)):
                    forms[j].instance.username = request.user
                    forms[j].instance.date = Organization.objects.order_by('-date')[0].date
                    forms[j].instance.num_week = j
                    forms[j].instance.M1 = request.POST.get(f"M1_{j}", False)
                    forms[j].instance.M2 = request.POST.get(f"M2_{j}", False)
                    forms[j].instance.M3 = request.POST.get(f"M3_{j}", False)
                    forms[j].instance.M4 = request.POST.get(f"M4_{j}", False)
                    forms[j].instance.M5 = request.POST.get(f"M5_{j}", False)
                    forms[j].instance.M6 = request.POST.get(f"M6_{j}", False)
                    forms[j].instance.M7 = request.POST.get(f"M7_{j}", False)
                    forms[j].instance.A1 = request.POST.get(f"A1_{j}", False)
                    forms[j].instance.A2 = request.POST.get(f"A2_{j}", False)
                    forms[j].instance.A3 = request.POST.get(f"A3_{j}", False)
                    forms[j].instance.A4 = request.POST.get(f"A4_{j}", False)
                    forms[j].instance.A5 = request.POST.get(f"A5_{j}", False)
                    forms[j].instance.A6 = request.POST.get(f"A6_{j}", False)
                    forms[j].instance.A7 = request.POST.get(f"A7_{j}", False)
                    forms[j].instance.N1 = request.POST.get(f"N1_{j}", False)
                    forms[j].instance.N2 = request.POST.get(f"N2_{j}", False)
                    forms[j].instance.N3 = request.POST.get(f"N3_{j}", False)
                    forms[j].instance.N4 = request.POST.get(f"N4_{j}", False)
                    forms[j].instance.N5 = request.POST.get(f"N5_{j}", False)
                    forms[j].instance.N6 = request.POST.get(f"N6_{j}", False)
                    forms[j].instance.N7 = request.POST.get(f"N7_{j}", False)
                    forms[j].instance.P1 = request.POST.get(f"P1_{j}", False)
                    forms[j].instance.P2 = request.POST.get(f"P2_{j}", False)
                    forms[j].instance.P3 = request.POST.get(f"P3_{j}", False)
                    forms[j].instance.P4 = request.POST.get(f"P4_{j}", False)
                    forms[j].instance.P5 = request.POST.get(f"P5_{j}", False)
                    forms[j].instance.P6 = request.POST.get(f"P6_{j}", False)
                    forms[j].instance.P7 = request.POST.get(f"P7_{j}", False)
                    forms[j].instance.R1 = request.POST.get(f"R1_{j}", False)
                    forms[j].instance.R2 = request.POST.get(f"R2_{j}", False)
                    forms[j].instance.R3 = request.POST.get(f"R3_{j}", False)
                    forms[j].instance.R4 = request.POST.get(f"R4_{j}", False)
                    forms[j].instance.R5 = request.POST.get(f"R5_{j}", False)
                    forms[j].instance.R6 = request.POST.get(f"R6_{j}", False)
                    forms[j].instance.R7 = request.POST.get(f"R7_{j}", False)
                    forms[j].instance.notes1 = request.POST.get(f"notes1_{j}", False)
                    forms[j].instance.notes2 = request.POST.get(f"notes2_{j}", False)
                    forms[j].instance.notes3 = request.POST.get(f"notes3_{j}", False)
                    forms[j].instance.notes4 = request.POST.get(f"notes4_{j}", False)
                    forms[j].instance.notes5 = request.POST.get(f"notes5_{j}", False)
                    forms[j].instance.notes6 = request.POST.get(f"notes6_{j}", False)
                    forms[j].instance.notes7 = request.POST.get(f"notes7_{j}", False)
                    forms[j].save()
            else:
                for j in range(organization.num_weeks):
                    weeks = ShiftWeek.objects.all().filter(date=organization.date).order_by('num_week')
                    new_shift = ShiftWeekForm(instance=weeks[j])
                    new_shift.username = request.user
                    new_shift.date = Organization.objects.order_by('-date')[0].date
                    new_shift.num_week = j
                    new_shift.M1 = request.POST.get(f"M1_{j}", False)
                    new_shift.M2 = request.POST.get(f"M2_{j}", False)
                    new_shift.M3 = request.POST.get(f"M3_{j}", False)
                    new_shift.M4 = request.POST.get(f"M4_{j}", False)
                    new_shift.M5 = request.POST.get(f"M5_{j}", False)
                    new_shift.M6 = request.POST.get(f"M6_{j}", False)
                    new_shift.M7 = request.POST.get(f"M7_{j}", False)
                    new_shift.A1 = request.POST.get(f"A1_{j}", False)
                    new_shift.A2 = request.POST.get(f"A2_{j}", False)
                    new_shift.A3 = request.POST.get(f"A3_{j}", False)
                    new_shift.A4 = request.POST.get(f"A4_{j}", False)
                    new_shift.A5 = request.POST.get(f"A5_{j}", False)
                    new_shift.A6 = request.POST.get(f"A6_{j}", False)
                    new_shift.A7 = request.POST.get(f"A7_{j}", False)
                    new_shift.N1 = request.POST.get(f"N1_{j}", False)
                    new_shift.N2 = request.POST.get(f"N2_{j}", False)
                    new_shift.N3 = request.POST.get(f"N3_{j}", False)
                    new_shift.N4 = request.POST.get(f"N4_{j}", False)
                    new_shift.N5 = request.POST.get(f"N5_{j}", False)
                    new_shift.N6 = request.POST.get(f"N6_{j}", False)
                    new_shift.N7 = request.POST.get(f"N7_{j}", False)
                    new_shift.P1 = request.POST.get(f"P1_{j}", False)
                    new_shift.P2 = request.POST.get(f"P2_{j}", False)
                    new_shift.P3 = request.POST.get(f"P3_{j}", False)
                    new_shift.P4 = request.POST.get(f"P4_{j}", False)
                    new_shift.P5 = request.POST.get(f"P5_{j}", False)
                    new_shift.P6 = request.POST.get(f"P6_{j}", False)
                    new_shift.P7 = request.POST.get(f"P7_{j}", False)
                    new_shift.R1 = request.POST.get(f"R1_{j}", False)
                    new_shift.R2 = request.POST.get(f"R2_{j}", False)
                    new_shift.R3 = request.POST.get(f"R3_{j}", False)
                    new_shift.R4 = request.POST.get(f"R4_{j}", False)
                    new_shift.R5 = request.POST.get(f"R5_{j}", False)
                    new_shift.R6 = request.POST.get(f"R6_{j}", False)
                    new_shift.R7 = request.POST.get(f"R7_{j}", False)
                    new_shift.notes1 = request.POST.get(f"notes1_{j}", False)
                    new_shift.notes2 = request.POST.get(f"notes2_{j}", False)
                    new_shift.notes3 = request.POST.get(f"notes3_{j}", False)
                    new_shift.notes4 = request.POST.get(f"notes4_{j}", False)
                    new_shift.notes5 = request.POST.get(f"notes5_{j}", False)
                    new_shift.notes6 = request.POST.get(f"notes6_{j}", False)
                    new_shift.notes7 = request.POST.get(f"notes7_{j}", False)
                    new_shift.save()
            if not already_submitted(request.user):
                messages.success(request, translate_text(f'משמרות הוגשו בהצלחה!', request.user, "hebrew"))
            else:
                messages.success(request, translate_text(f'משמרות עודכנו בהצלחה!', request.user, "hebrew"))
            return redirect("Schedule-Home")
        else:
            messages.error(request, translate_text(f'שינויים לא נשמרו!', request.user, "hebrew"))
    else:
        if not submitting:
            shifts = Shift.objects.order_by('-date')
            if len(shifts.filter(username=request.user, date=organization.date).order_by('-date')) > 0:
                shift = shifts.filter(username=request.user).order_by('-date')[0]
                weeks = shifts_weeks_served.filter(username=request.user).order_by('num_week')
                notes_text = str(shift.notes)
                for i in range(len(weeks)):
                    forms[i] = ShiftWeekViewForm(instance=weeks[i])
                #form = ShiftViewForm(instance=shift)
            else:
                empty = True
        elif not already_submitted(request.user):
            #form = ShiftForm()
            shift = Shift()
            for i in range(organization.num_weeks):
                forms[i] = ShiftWeekForm()
        else:
            last_date = Organization.objects.order_by('-date')[0].date
            shifts = Shift.objects.filter(date=last_date)
            shift = shifts.filter(username=request.user).first()
            notes_text = str(shift.notes)
            #form = ShiftForm(instance=shift)
            weeks = shifts_weeks_served.filter(username=request.user).order_by('num_week')
            for i in range(len(weeks)):
                forms[i] = ShiftWeekForm(instance=weeks[i])
    if not empty:
        context = {
            "form": shift,
            "days": days,
            "submitting": submitting,
            "notes_text": notes_text,
            "empty": empty,
            "manager": False,
            "forms": forms,
        }
    else:
        context = {
            "empty": empty,
            "manager": False,
        }
    return render(request, "Schedule/shifts.html", context)


def already_submitted(user):
    last_date = Organization.objects.order_by('-date')[0].date
    shifts = Shift.objects.filter(date=last_date)
    if len(shifts) == 0:
        return False
    else:
        if len(shifts.filter(username=user)) == 0:
            return False
    return True


def get_input(organization_last):
    weeks_obj = Week.objects.all().filter(date=organization_last.date)
    weeks = []
    for w in weeks_obj:
        weeks.append(0)
    for week_obj in weeks_obj:
        weeks[week_obj.num_week] = week_obj
    organization_last_input = {}
    fields = ["_630", "_700_search", "_700_manager", "_720_1", "_720_pull", "_720_2", "_720_3", "_1400", "_1500",
              "_1500_1900", "_2300", "_notes"]
    for j in range(organization_last.num_weeks):
        for i in range(1, 8):
            day1 = f'Day{i}'
            day2 = f'day{i + (j * 7)}'
            for f in fields:
                organization_last_input[day2 + f] = getattr(weeks[j], day1 + f)
    return organization_last_input


class ServedSumListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Organization
    template_name = "Schedule/Served_sum_list.html"
    context_object_name = "organizations"
    ordering = ["-date"]
    paginate_by = 5

    def test_func(self):
        if self.request.user.is_staff:
            return True
        return False


@staff_member_required
def shift_update_view(request, pk=None):
    main_shift = Shift.objects.all().filter(id=pk).first()
    organization = Organization.objects.all().filter(date=main_shift.date).first()
    shifts_weeks_served = ShiftWeek.objects.all().filter(date=organization.date)
    user = User.objects.filter(username=main_shift.username).first()
    forms = []
    for i in range(organization.num_weeks):
        forms.append("")
    form = None
    notes_text = ""
    user_settings = USettings.objects.all().filter(user=user).first()
    days = []
    for x in range(organization.num_weeks * 7):
        days.append(organization.date + datetime.timedelta(days=x))
    events = Event.objects.all()
    for x in range(organization.num_weeks * 7):
        if len(events.filter(date2=days[x])) > 0:
            for ev in events.filter(date2=days[x]):
                if user_settings.nickname == ev.nickname:
                    message = f'לא לשכוח בתאריך {ev.date2} יש {ev.description}. אם יש שינוי להודיע!'
                    message = translate_text(message, user, "hebrew")
                    messages.info(request, message)
                elif ev.nickname == 'כולם':
                    message = f'לא לשכוח בתאריך {ev.date2} יש {ev.description}'
                    message = translate_text(message, user, "hebrew")
                    messages.info(request, message)
    if request.method == 'POST':
        last_date = organization.date
        shifts = Shift.objects.filter(date=last_date)
        shift = shifts.filter(username=user).first()
        notes_text = str(shift.notes)
        weeks = shifts_weeks_served.filter(username=user).order_by('num_week')
        for week in weeks:
            new_form = ShiftWeekForm(request.POST, instance=week)
            forms[week.num_week] = new_form
        shift.username = user
        shift.date = organization.date
        notes_area = request.POST.get("notesArea")
        shift.notes = notes_area
        shift.seq_night = request.POST.get(f"seq_night", 0)
        shift.seq_noon = request.POST.get(f"seq_noon", 0)
        error = False
        shift.save()
        for form in forms:
            if not form.is_valid():
                error = True
        if not error:
            for j in range(len(forms)):
                forms[j].instance.username = user
                forms[j].instance.date = organization.date
                forms[j].instance.num_week = j
                forms[j].instance.M1 = request.POST.get(f"M1_{j}", False)
                forms[j].instance.M2 = request.POST.get(f"M2_{j}", False)
                forms[j].instance.M3 = request.POST.get(f"M3_{j}", False)
                forms[j].instance.M4 = request.POST.get(f"M4_{j}", False)
                forms[j].instance.M5 = request.POST.get(f"M5_{j}", False)
                forms[j].instance.M6 = request.POST.get(f"M6_{j}", False)
                forms[j].instance.M7 = request.POST.get(f"M7_{j}", False)
                forms[j].instance.A1 = request.POST.get(f"A1_{j}", False)
                forms[j].instance.A2 = request.POST.get(f"A2_{j}", False)
                forms[j].instance.A3 = request.POST.get(f"A3_{j}", False)
                forms[j].instance.A4 = request.POST.get(f"A4_{j}", False)
                forms[j].instance.A5 = request.POST.get(f"A5_{j}", False)
                forms[j].instance.A6 = request.POST.get(f"A6_{j}", False)
                forms[j].instance.A7 = request.POST.get(f"A7_{j}", False)
                forms[j].instance.N1 = request.POST.get(f"N1_{j}", False)
                forms[j].instance.N2 = request.POST.get(f"N2_{j}", False)
                forms[j].instance.N3 = request.POST.get(f"N3_{j}", False)
                forms[j].instance.N4 = request.POST.get(f"N4_{j}", False)
                forms[j].instance.N5 = request.POST.get(f"N5_{j}", False)
                forms[j].instance.N6 = request.POST.get(f"N6_{j}", False)
                forms[j].instance.N7 = request.POST.get(f"N7_{j}", False)
                forms[j].instance.P1 = request.POST.get(f"P1_{j}", False)
                forms[j].instance.P2 = request.POST.get(f"P2_{j}", False)
                forms[j].instance.P3 = request.POST.get(f"P3_{j}", False)
                forms[j].instance.P4 = request.POST.get(f"P4_{j}", False)
                forms[j].instance.P5 = request.POST.get(f"P5_{j}", False)
                forms[j].instance.P6 = request.POST.get(f"P6_{j}", False)
                forms[j].instance.P7 = request.POST.get(f"P7_{j}", False)
                forms[j].instance.R1 = request.POST.get(f"R1_{j}", False)
                forms[j].instance.R2 = request.POST.get(f"R2_{j}", False)
                forms[j].instance.R3 = request.POST.get(f"R3_{j}", False)
                forms[j].instance.R4 = request.POST.get(f"R4_{j}", False)
                forms[j].instance.R5 = request.POST.get(f"R5_{j}", False)
                forms[j].instance.R6 = request.POST.get(f"R6_{j}", False)
                forms[j].instance.R7 = request.POST.get(f"R7_{j}", False)
                forms[j].instance.notes1 = request.POST.get(f"notes1_{j}", False)
                forms[j].instance.notes2 = request.POST.get(f"notes2_{j}", False)
                forms[j].instance.notes3 = request.POST.get(f"notes3_{j}", False)
                forms[j].instance.notes4 = request.POST.get(f"notes4_{j}", False)
                forms[j].instance.notes5 = request.POST.get(f"notes5_{j}", False)
                forms[j].instance.notes6 = request.POST.get(f"notes6_{j}", False)
                forms[j].instance.notes7 = request.POST.get(f"notes7_{j}", False)
                forms[j].save()
            messages.success(request, translate_text(f'משמרות עודכנו בהצלחה!', user, "hebrew"))
            return redirect("Schedule-Served-sum")
        else:
            messages.error(request, translate_text(f'שינויים לא נשמרו!', user, "hebrew"))
    else:
        last_date = organization.date
        shifts = Shift.objects.filter(date=last_date)
        shift = shifts.filter(username=user).first()
        notes_text = str(shift.notes)
        # form = ShiftForm(instance=shift)
        weeks = shifts_weeks_served.filter(username=user).order_by('num_week')
        for i in range(len(weeks)):
            forms[i] = ShiftWeekForm(instance=weeks[i])
    context = {
            "form": shift,
            "days": days,
            "submitting": True,
            "notes_text": notes_text,
            "empty": False,
            "manager": True,
            "forms": forms,
            "userview": USettings.objects.all().filter(user=user).first().nickname,
    }
    return render(request, "Schedule/shifts.html", context)


class OrganizationDetailView(LoginRequiredMixin, DetailView):
    model = Organization
    template_name = "Schedule/organization-detail.html"

    def get_context_data(self, **kwargs):
        ctx = super(OrganizationDetailView, self).get_context_data(**kwargs)
        weeks_obj = Week.objects.all().filter(date=self.get_object().date)
        weeks = []
        for w in weeks_obj:
            weeks.append(0)
        for week_obj in weeks_obj:
            weeks[week_obj.num_week] = week_obj
        ctx["weeks"] = weeks
        days = []
        for x in range(self.get_object().num_weeks * 7):
            days.append(self.get_object().date + datetime.timedelta(days=x))
        ctx["days"] = days
        return ctx


class OrganizationCreateView(LoginRequiredMixin, CreateView, UserPassesTestMixin):
    model = Organization
    template_name = "Schedule/organization-new.html"
    fields = ('date', 'num_weeks')

    def get_context_data(self, **kwargs):
        ctx = super(OrganizationCreateView, self).get_context_data(**kwargs)
        ctx["date1"] = timezone.now()
        return ctx

    def test_func(self):
        if self.request.user.is_staff:
            return True
        return False

    def form_valid(self, form):
        form.instance.date = self.request.POST.get("date")
        return super().form_valid(form)


class ShifttableView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Organization
    template_name = "Schedule/shift-table-view.html"

    def get_context_data(self, **kwargs):
        ctx = super(ShifttableView, self).get_context_data(**kwargs)
        organization = get_input(self.get_object())
        users = set()
        for key in organization:
            if key.count("notes") == 0:
                organization[key] = organization[key].replace("\r", "\n")
                split = organization[key].split("\n")
                for s in split:
                    s = s.replace(" ", "")
                    users.add(s)
        if " " in users:
            users.remove(" ")
        if "" in users:
            users.remove("")
        table_content = {}
        sum_content = {}
        for i in range(self.get_object().num_weeks):
            sum_content[f'morning{i + 1}'] = 0
            sum_content[f'noon{i + 1}'] = 0
        sum_content["night"] = 0
        sum_content["end"] = 0
        for user in users:
            table_content[user] = {}
            for i in range(self.get_object().num_weeks):
                table_content[user][f'morning{i + 1}'] = 0
                table_content[user][f'noon{i + 1}'] = 0
            table_content[user]["night"] = 0
            table_content[user]["end"] = 0
        morning_shifts = ["630", "700_search", "700_manager", "720_pull", "720_1", "720_2", "720_3"]
        noon_shifts = ["1400", "1500", "1500_1900"]
        num_week = 0
        count = 0
        for i in range(1, self.get_object().num_weeks * 7 + 1):
            count += 1
            day = f'day{i}_'
            for shift in morning_shifts:
                split = organization[f'{day}{shift}'].split("\n")
                for s in split:
                    s = s.replace(" ", "")
                    if s != "":
                        if count != 7:
                            table_content[s][f"morning{num_week + 1}"] += 1
                            sum_content[f"morning{num_week + 1}"] += 1
                        else:
                            table_content[s]["end"] += 1
                            sum_content["end"] += 1
            for shift in noon_shifts:
                split = organization[f'{day}{shift}'].split("\n")
                for s in split:
                    s = s.replace(" ", "")
                    if s != "":
                        if count < 6:
                            table_content[s][f"noon{num_week + 1}"] += 1
                            sum_content[f"noon{num_week + 1}"] += 1
                        else:
                            table_content[s]["end"] += 1
                            sum_content["end"] += 1
            split = organization[f'{day}2300'].split("\n")
            for s in split:
                s = s.replace(" ", "")
                if s != "":
                    if count < 6:
                        table_content[s]["night"] += 1
                        sum_content["night"] += 1
                    else:
                        table_content[s]["end"] += 1
                        sum_content["end"] += 1
            if count == 7:
                count = 0
                num_week += 1
        print(table_content)
        ctx["table"] = table_content
        ctx["sum"] = sum_content
        days = []
        for x in range(self.get_object().num_weeks * 7):
            days.append(self.get_object().date + datetime.timedelta(days=x))
        ctx["days"] = days
        ctx["num_weeks"] = self.get_object().num_weeks + 1
        return ctx

    def test_func(self):
        if self.request.user.is_staff:
            return True
        return False


class ServedSumReinforcementsDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Organization
    template_name = "Schedule/served_sum_reinforcements.html"

    def get_data(self, calculated: bool):
        ctx = {}
        served = {}
        if calculated:
            organization = self.get_object()
            organization_input = get_input(organization)
            input_days = {}
            keys = ["_630", "_700_search", "_720_1", "_720_pull", "_720_2", "_720_3", "_1400", "_1500",
                    "_1500_1900",
                    "_2300"]
        for i in range(1, 15):
            served["day" + str(i)] = ""
            if calculated:
                day = "day" + str(i)
                input_days[day + "M"] = []
                input_days[day + "A"] = []
                input_days[day + "N"] = []
                for x in range(10):
                    if x < 6:
                        input_days[day + "M"] += organization_input[day + keys[x]].split("\n")
                    elif x < 9:
                        input_days[day + "A"] += organization_input[day + keys[x]].split("\n")
                    else:
                        input_days[day + "N"] += organization_input[day + keys[x]].split("\n")
        if calculated:
            for key in input_days:
                for i in range(len(input_days[key])):
                    input_days[key][i] = input_days[key][i].replace(" ", "")
                    input_days[key][i] = input_days[key][i].replace("\n", "")
                    input_days[key][i] = input_days[key][i].replace("\r", "")
        shift_date = self.get_object().date
        main_shifts_served = Shift.objects.all().filter(date=shift_date)
        shifts_served = ShiftWeek.objects.all().filter(date=shift_date)
        weeks_notes = []
        for i in range(self.get_object().num_weeks):
            weeks_notes.append("")
        notes_general = ""
        users = {}
        user_notes_added = []
        for shift in shifts_served:
            username = shift.username
            user = User.objects.all().filter(username=username).first()
            user_settings = USettings.objects.all().filter(user=user).first()
            main_shift = main_shifts_served.filter(username=username).first()
            users[user_settings.nickname] = main_shifts_served.filter(username=username).first().id
            name = user_settings.nickname
            index = 1
            shifts = [shift.R1, shift.R2, shift.R3, shift.R4, shift.R5, shift.R6, shift.R7]
            for s in shifts:
                if s:
                    served["day" + str(index + (shift.num_week * 7))] += name
                index += 1
            notes = [shift.notes1, shift.notes2, shift.notes3,
                      shift.notes4, shift.notes5, shift.notes6, shift.notes7]
            index = 1
            for n in notes:
                if n != "":
                    notes[shift.num_week] += name + ": " \
                                     + number_to_day2(index) + " - " + n + "\n"
                index += 1
            if main_shift.notes != "" and name not in user_notes_added:
                notes_general += name + ": " + main_shift.notes + "\n"
                user_notes_added.append(name)
        days = []
        for x in range(self.get_object().num_weeks * 7):
            days.append(self.get_object().date + datetime.timedelta(days=x))
        ##
        # Calculated Part
        if calculated:
            calc_served = {}
            for x in range(1, self.get_object().num_weeks * 7 + 1):
                day = "day" + str(x)
                calc_served[day] = ""
                split = served[day].split("\n")
                for s in split:
                    if s not in input_days[day + "A"] and s != "" and s != " " and s != "\n":
                        day_before = "day" + str(x - 1)
                        day_after = "day" + str(x + 1)
                        if x != 1 and x != 14:
                            if s not in input_days[day + "M"] and s not in input_days[day + "N"] and \
                                    s not in input_days[day_before + "N"] and s not in input_days[day_after + "M"]:
                                calc_served[day] += s + "\n"
                            elif s not in input_days[day + "M"] and s not in input_days[day_before + "N"] \
                                    and s not in input_days[day + "N"]:
                                calc_served[day] += s + "\n" + " (יכול בוקר וצהריים) " + "\n"
                            elif s not in input_days[day + "M"] and s not in input_days[day_before + "N"]:
                                calc_served[day] += s + "\n" + " (יכול רק בוקר) " + "\n"
                            elif s not in input_days[day + "M"] and s not in input_days[day_after + "M"] \
                                    and s not in input_days[day + "N"]:
                                calc_served[day] += s + "\n" + " (יכול רק צהריים ולילה) " + "\n"
                            elif s not in input_days[day + "N"] and s not in input_days[day_after + "M"]:
                                calc_served[day] += s + "\n" + " (יכול רק לילה) " + "\n"
                        elif x == 1:
                            if s not in input_days[day + "M"] and s not in input_days[day + "N"] and \
                                    s not in input_days[day_after + "M"]:
                                calc_served[day] += s + "\n"
                            elif s not in input_days[day + "M"] and s not in input_days[day + "N"]:
                                calc_served[day] += s + "\n" + " (יכול בוקר וצהריים) " + "\n"
                            elif s not in input_days[day + "M"] and s not in input_days[day_after + "M"] \
                                    and s not in input_days[day + "N"]:
                                calc_served[day] += s + "\n" + " (יכול רק צהריים ולילה) " + "\n"
                            elif s not in input_days[day + "N"] and s not in input_days[day_after + "M"]:
                                calc_served[day] += s + "\n" + " (יכול רק לילה) " + "\n"
                            elif s not in input_days[day + "M"]:
                                calc_served[day] += s + "\n" + " (יכול רק בוקר) " + "\n"
                        else:
                            if s not in input_days[day + "M"] and s not in input_days[day + "N"] and \
                                    s not in input_days[day_before + "N"]:
                                calc_served[day] += s + "\n"
                            elif s not in input_days[day + "M"] and s not in input_days[day_before + "N"] \
                                    and s not in input_days[day + "N"]:
                                calc_served[day] += s + "\n" + " (יכול בוקר וצהריים) " + "\n"
                            elif s not in input_days[day + "M"] and s not in input_days[day_before + "N"]:
                                calc_served[day] += s + "\n" + " (יכול רק בוקר) " + "\n"
                            elif s not in input_days[day + "M"] and s not in input_days[day + "N"]:
                                calc_served[day] += s + "\n" + " (יכול רק צהריים ולילה) " + "\n"
                            elif s not in input_days[day + "N"]:
                                calc_served[day] += s + "\n" + " (יכול רק לילה) " + "\n"
            served = calc_served
        ctx["calculated"] = calculated
        ctx["days"] = days
        ctx["served"] = served
        ctx["notes"] = weeks_notes
        ctx["notes_general"] = notes_general
        ctx["num_served"] = len(shifts_served)
        ctx["users"] = users
        return ctx

    def get_context_data(self, **kwargs):
        ctx = super(ServedSumReinforcementsDetailView, self).get_context_data(**kwargs)
        calculated = False
        context = self.get_data(calculated)
        for c in context:
            ctx[c] = context[c]
        return ctx

    def test_func(self):
        if self.request.user.is_staff:
            return True
        return False

    def post(self, request, *args, **kwargs):
        if request.method == "POST":
            calculated = True
            if 'calculated' in request.POST:
                if request.POST.get("calculated") == "True":
                    calculated = False
                else:
                    calculated = True
            ctx = self.get_data(calculated)
            return render(request, "Schedule/served_sum_reinforcements.html", ctx)


class ServedSumShiftDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Organization
    template_name = "Schedule/Served-sum.html"

    def get_data(self):
        ctx = {}
        served = {}
        for i in range(1, self.get_object().num_weeks * 7 + 1):
            served["M" + str(i)] = ""
            served["A" + str(i)] = ""
            served["N" + str(i)] = ""
        shift_date = self.get_object().date
        main_shifts_served = Shift.objects.all().filter(date=shift_date)
        shifts_served = ShiftWeek.objects.all().filter(date=shift_date)
        weeks_notes = []
        for i in range(self.get_object().num_weeks):
            weeks_notes.append("")
        notes_general = ""
        user_notes_added = []
        users = {}
        for shift in shifts_served:
            username = shift.username
            user = User.objects.all().filter(username=username).first()
            user_settings = USettings.objects.all().filter(user=user).first()
            main_shift = main_shifts_served.filter(username=username).first()
            users[user_settings.nickname] = main_shift.id
            name = user_settings.nickname
            name = name.replace("\n", "")
            name = name.replace("\r", "")
            kind = "M"
            morning = False
            count = 0
            index = 1
            shifts = [shift.M1, shift.P1, shift.A1, shift.N1, shift.M2, shift.P2, shift.A2, shift.N2, shift.M3,
                      shift.P3,
                      shift.A3, shift.N3, shift.M4, shift.P4, shift.A4, shift.N4, shift.M5, shift.P5, shift.A5,
                      shift.N5,
                      shift.M6, shift.P6, shift.A6, shift.N6, shift.M7, shift.P7, shift.A7, shift.N7]
            for s in shifts:
                if s:
                    if count == 0:
                        morning = True
                        served[kind + str(index + (shift.num_week * 7))] += name
                    else:
                        if count == 1:
                            if morning:
                                served[kind + str(index + (shift.num_week * 7))] += "\n"
                            morning = False
                        else:
                            if count == 2:
                                kind = "A"
                            elif count == 3:
                                kind = "N"
                            served[kind + str(index + (shift.num_week * 7))] += name + "\n"
                else:
                    if count == 1 and morning:
                        morning = False
                        served[kind + str(index + (shift.num_week * 7))] += "\n" + "(לא משיכה)" + "\n"
                if count == 3:
                    count = 0
                    index = index + 1
                    kind = "M"
                else:
                    count = count + 1
            notes1 = [shift.notes1, shift.notes2, shift.notes3,
                      shift.notes4, shift.notes5, shift.notes6, shift.notes7]
            index = 1
            for n in notes1:
                if n != "":
                    weeks_notes[shift.num_week] += name + ": " \
                                     + number_to_day2(index) + " - " + n + "\n"
                index += 1
            if main_shift.notes != "" and name not in user_notes_added:
                notes_general += name + ": " + main_shift.notes + "\n"
                user_notes_added.append(name)
        days = []
        for x in range(self.get_object().num_weeks * 7):
            days.append(self.get_object().date + datetime.timedelta(days=x))
        ctx["days"] = days
        ctx["served"] = served
        ctx["notes"] = weeks_notes
        ctx["notes_general"] = notes_general
        ctx["num_served"] = len(main_shifts_served)
        ctx["users"] = users
        return ctx

    def get_context_data(self, **kwargs):
        ctx = super(ServedSumShiftDetailView, self).get_context_data(**kwargs)
        context = self.get_data()
        for c in context:
            ctx[c] = context[c]
        return ctx

    def test_func(self):
        if self.request.user.is_staff:
            return True
        return False

    def post(self, request, *args, **kwargs):
        if request.method == "POST":
            ctx = self.get_data()
            return WriteToExcel(ctx["served"], ctx["notes"], ctx["notes_general"],ctx["days"], self.request.user)


@staff_member_required
def organization_update(request, pk=None):
    organization = Organization.objects.all().filter(id=pk).first()
    weeks = Week.objects.all().filter(date=organization.date)
    days = []
    for x in range(organization.num_weeks * 7):
        days.append(organization.date + datetime.timedelta(days=x))
    if request.method == "POST":
        action = request.POST.get("actions")
        forms = []
        for i in range(organization.num_weeks):
            forms.append("")
        for week in weeks:
            new_form = WeekUpdateForm(request.POST, instance=week)
            forms[week.num_week] = new_form
        error = False
        for form in forms:
            if not form.is_valid():
                error = True
        if not error:
            for j in range(len(forms)):
                to_week_form(forms[j], request, j)
            for form in forms:
                form.save()
            pub = request.POST.get("pub")
            if pub:
                published = False
            else:
                published = True
            organization.published = published
            organization.save()
            messages.success(request, translate_text(f'עדכון הושלם', request.user, "hebrew"))
            if 'update' == action:
                return HttpResponseRedirect(request.path_info)
        else:
            messages.info(request, translate_text(f'תקלה טכנית לא ניתן לבדוק', request.user, "hebrew"))
            return HttpResponseRedirect(request.path_info)
        if 'check1' == action:
            organization_valid(organization, request)
            return HttpResponseRedirect(request.path_info)
        elif 'ready' == action:
            return redirect("organization-detail", organization.id)
        elif 'table' == action:
            return redirect("organization-table-shift", organization.id)
        elif 'upload' == action:
            weeks_dicts = uplaod_organize(request, organization)
            forms = []
            for i in range(organization.num_weeks):
                forms.append("")
            #for week in weeks:
            #    new_form = WeekUpdateForm(request.POST, instance=week)
            #    forms[week.num_week] = new_form
            for i in range(len(weeks)):
                for key in weeks_dicts[i].keys():
                    setattr(weeks[i], key, weeks_dicts[i][key])
            error = False
            #for form in forms:
            #    if not form.is_valid():
            #        error = True
            if not error:
                for week in weeks:
                    week.save()
                messages.success(request, translate_text(f'העלאה הושלמה', request.user, "hebrew"))
            else:
                messages.info(request, translate_text(f'עדכון לא הושלם תקלה טכנית', request.user, "hebrew"))
            return HttpResponseRedirect(request.path_info)
        elif 'clear' == action:
            forms = []
            for i in range(organization.num_weeks):
                forms.append("")
            for week in weeks:
                new_form = WeekUpdateForm(request.POST, instance=week)
                forms[week.num_week] = new_form
            for j in range(len(forms)):
                forms[j].data._mutable = True
                for i in range(1, 8):
                    forms[j].data["Day" + str(i) + "_630"] = ""
                    forms[j].data["Day" + str(i) + "_700_search"] = ""
                    forms[j].data["Day" + str(i) + "_700_manager"] = ""
                    forms[j].data["Day" + str(i) + "_720_1"] = ""
                    forms[j].data["Day" + str(i) + "_720_pull"] = ""
                    forms[j].data["Day" + str(i) + "_720_2"] = ""
                    forms[j].data["Day" + str(i) + "_720_3"] = ""
                    forms[j].data["Day" + str(i) + "_1400"] = ""
                    forms[j].data["Day" + str(i) + "_1500"] = ""
                    forms[j].data["Day" + str(i) + "_1500_1900"] = ""
                    forms[j].data["Day" + str(i) + "_2300"] = ""
                forms[j].data._mutable = False
            error = False
            for form in forms:
                if not form.is_valid():
                    error = True
            if not error:
                for form in forms:
                    form.save()
                messages.success(request, translate_text(f'איפוס הושלם', request.user, "hebrew"))
            else:
                messages.info(request, translate_text(f'איפוס לא הושלם תקלה טכנית', request.user, "hebrew"))
            return HttpResponseRedirect(request.path_info)
        elif 'delete' == action:
            Organization.objects.filter(id=organization.id).delete()
            messages.success(request, translate_text(f'מחיקה הושלמה', request.user, "hebrew"))
            return redirect("Schedule-Served-sum")
    else:
        forms = []
        for i in range(organization.num_weeks):
            forms.append("")
        for week in weeks:
            new_form = WeekUpdateForm(instance=week)
            forms[week.num_week] = new_form
    context = {
        "forms": forms,
        "checked": organization.published,
        "organization_id": organization.id,
        "days": days,
    }
    return render(request, "Schedule/organization_update.html", context)


def extract_data(request, organization):
    # extract from excel
    myfile = request.FILES['myfile']
    file_object = myfile.file
    wb = openpyxl.load_workbook(file_object)
    sheet = wb.active
    ## Green background FFC6EFCE
    ## Green font FF006100
    ## Red background FFFFC7CE
    ## Red font with red background FF9C0006
    ## White background 00000000
    ## Black font Values must be of type <class 'str'> rgb=None, indexed=None, auto=None, theme=1, tint=0.0, type='theme'
    ## Red font FFFF0000
    ## Empty font Values must be of type <class 'str'> rgb=None, indexed=None, auto=None, theme=1, tint=0.0, type='theme'
    ## Empty value None
    ## orange background FFFFEB9C
    # print(str(sheet.cell(10, 3).value))
    # print(str(sheet.cell(10, 3).font.color.rgb))  # Get the font color in the table
    # print(str(sheet.cell(10, 3).fill.fgColor.rgb)) # background color
    # if str(sheet.cell(10, 3).font.color.rgb) == "Values must be of type <class 'str'>":
    end_morning_str = ""
    end_noon_str = ""
    end_night_str = ""
    for rng in sheet.merged_cells.ranges:
        if 'A5' in rng:
            end_morning_str = str(rng)
            break
    end_morning_str = end_morning_str.replace("A5:A", "")
    end_morning = int(end_morning_str)
    for rng in sheet.merged_cells.ranges:
        if f'A{end_morning + 1}' in rng:
            end_noon_str = str(rng)
            break
    end_noon_str = end_noon_str.replace(f'A{end_morning + 1}:A', "")
    end_noon = int(end_noon_str)
    for rng in sheet.merged_cells.ranges:
        if f'A{end_noon + 1}' in rng:
            end_night_str = str(rng)
            break
    end_night_str = end_night_str.replace(f'A{end_noon + 1}:A', "")
    end_night = int(end_night_str)
    col = 1
    names_days = {}
    no_pull_names = {}
    for x in range(14):
        col += 1
        names_days[f'day{x}_morning'] = []
        no_pull_names[f'day{x}'] = []
        names_days[f'day{x}_noon'] = []
        names_days[f'day{x}_night'] = []
        for j in range(5, end_morning + 1):
            if str(sheet.cell(j, col).fill.fgColor.rgb) == 'FFC6EFCE' \
                    or str(sheet.cell(j, col).fill.fgColor.rgb) == 'FFFFEB9C':
                names_days[f'day{x}_morning'].append(str(sheet.cell(j, col).value))
                if str(sheet.cell(j, col).fill.fgColor.rgb) == 'FFFFEB9C':
                    no_pull_names[f'day{x}'].append(str(sheet.cell(j, col).value))
        for j in range(end_morning + 1, end_noon + 1):
            if str(sheet.cell(j, col).fill.fgColor.rgb) == 'FFC6EFCE':
                names_days[f'day{x}_noon'].append(str(sheet.cell(j, col).value))
        for j in range(end_noon + 1, end_night + 1):
            if str(sheet.cell(j, col).fill.fgColor.rgb) == 'FFC6EFCE':
                names_days[f'day{x}_night'].append(str(sheet.cell(j, col).value))
    # extract from database
    shifts = Shift.objects.all().filter(date=organization.date)
    users = User.objects.all()
    settings = Settings.objects.all().first()
    max_seq0 = settings.max_seq0
    max_seq1 = settings.max_seq1
    sequence_count = {}
    max_out_names = [[], []]
    for s in shifts:
        user = users.filter(username=s.username).first()
        user_settings = USettings.objects.all().filter(user=user).first()
        name = user_settings.nickname
        sequence_count[f'{name}0'] = s.seq_night
        sequence_count[f'{name}1'] = s.seq_noon
        if sequence_count[f'{name}0'] >= max_seq0:
            max_out_names[0].append(name)
        if sequence_count[f'{name}1'] >= max_seq1:
            max_out_names[1].append(name)
    return [names_days, no_pull_names, sequence_count, max_out_names, max_seq0, max_seq1]


def uplaod_organize(request, organization):
    extracted_data = extract_data(request, organization)
    names_days = extracted_data[0]
    no_pull_names = extracted_data[1]
    sequence_count = extracted_data[2]
    max_out_names = extracted_data[3]
    max_seq0 = extracted_data[4]
    max_seq1 = extracted_data[5]
    try:
        before_organization = Organization.objects.order_by('-date')[1]
    except:
        before_organization = None
    if before_organization is not None:
        week_before = Week.objects.all().filter(date=before_organization.date,
                                                num_week=before_organization.num_weeks - 1).first()
        before_names = {"motsash": week_before.Day7_2300.split("\n"),
                        "noon": week_before.Day7_1500.split("\n"),
                        "morning": week_before.Day7_630.split("\n") +
                                   week_before.Day7_700_search.split("\n") +
                                   week_before.Day7_700_manager.split("\n") +
                                   week_before.Day7_720_1.split("\n")}
        for key in before_names:
            for v in before_names[key]:
                if v == '' or v == '\r' or v == ' ':
                    count = before_names[key].count(v)
                    for x in range(count):
                        before_names[key].remove(v)
                else:
                    before_names[key][before_names[key].index(v)] = v.replace("\r", "")
    else:
        before_names = {"motsash": [], "noon": [], "morning": []}
    weeks = Week.objects.all().filter(date=organization.date)
    manager_group = Group.objects.filter(name="manager").first()
    manager_group_users = User.objects.filter(groups=manager_group)
    managers = []
    for m in manager_group_users:
        user_settings = USettings.objects.all().filter(user=m).first()
        managers.append(user_settings.nickname)
    num_week = organization.num_weeks - 1
    days_count = 0
    weeks_dicts = []
    for i in range(organization.num_weeks):
        weeks_dicts.append({})
    no_fields = ["id", "date", "num_week"]
    for i in range(organization.num_weeks):
        fields = model_to_dict(weeks[0]).keys()
        for field in fields:
            if field not in no_fields:
                weeks_dicts[i][field] = ""
    for i in range(organization.num_weeks * 7 - 1, -1, -1):
        names_x = i
        x = i - 7 * num_week + 1
        if x != 6 and x != 7:
            is_manager = False
            for name in managers:
                if name in names_days[f'day{names_x}_morning']:
                    weeks_dicts[num_week][f'Day{x}_700_manager'] = name
                    names_days[f'day{names_x}_morning'].remove(name)
                    is_manager = True
                    break
            if not is_manager:
                if Settings.objects.last().officer in names_days[f'day{names_x}_morning']:
                    weeks_dicts[num_week][f'Day{x}_700_manager'] = Settings.objects.last().officer
                    names_days[f'day{names_x}_morning'].remove(Settings.objects.last().officer)
            temp_morning = []
            if len(names_days[f'day{names_x}_morning']) > 0:
                for name in names_days[f'day{names_x}_morning']:
                    temp_morning.append(name)
                for name in temp_morning:
                    if name in no_pull_names[f'day{names_x}'] or\
                            name in names_days[f'day{names_x}_night']:
                        if x == 1:
                            temp_morning.remove(name)
                        elif name in names_days[f'day{names_x - 1}_noon']:
                            temp_morning.remove(name)
                if len(temp_morning) > 0:
                    inserted = insert_random(weeks_dicts[num_week], temp_morning, "720_pull", x, 0)
                    names_days[f'day{names_x}_morning'].remove(inserted)
                else:
                    temp_morning = []
                    for name in names_days[f'day{names_x}_morning']:
                        temp_morning.append(name)
                    for name in temp_morning:
                        if x != 1:
                            if name in names_days[f'day{names_x - 1}_noon']:
                                temp_morning.remove(name)
                        else:
                            if name in before_names["noon"] or name in before_names["morning"]:
                                temp_morning.remove(name)
                    if len(temp_morning) > 0:
                        inserted = insert_random(weeks_dicts[num_week], temp_morning, "720_pull", x, 0)
                        names_days[f'day{names_x}_morning'].remove(inserted)
                    else:
                        insert_random(weeks_dicts[num_week], names_days[f'day{names_x}_morning'], "720_pull", x, 0)
            chosen = False
            if x == 1 and num_week == 0:
                chosen = search_and_put(weeks_dicts[num_week], before_names["noon"], names_days[f'day{names_x}_morning'], x,
                                             "630", max_out_names[0], 0, sequence_count, max_seq0,
                                             max_seq1, True, [])
                if not chosen:
                    chosen = search_and_put(weeks_dicts[num_week], before_names["morning"], names_days[f'day{names_x}_morning'], x,
                                                 "630", max_out_names[0], 0, sequence_count, max_seq0,
                                                 max_seq1, True, [])
                if not chosen:
                    temp_morning = seperate_list(names_days[f'day{names_x}_morning'], max_out_names)
                    if len(temp_morning) > 0:
                        chosen = insert_random(weeks_dicts[num_week], temp_morning, "630", x, 0)
                        if chosen is not None:
                            names_days[f'day{names_x}_morning'].remove(chosen)
                    else:
                        insert_random(weeks_dicts[num_week], names_days[f'day{names_x}_morning'], "630", x, 0)
                chosen = search_and_put(weeks_dicts[num_week], before_names["noon"], names_days[f'day{names_x}_morning'], x,
                                             "700_search", max_out_names[0], 0, sequence_count, max_seq0,
                                             max_seq1, False, [])
                if not chosen:
                    chosen = search_and_put(weeks_dicts[num_week], before_names["morning"], names_days[f'day{names_x}_morning'], x,
                                                 "700_search", max_out_names[0], 0, sequence_count, max_seq0,
                                                 max_seq1, False, [])
                if not chosen:
                    insert_random(weeks_dicts[num_week], names_days[f'day{names_x}_morning'], "700_search", x, 0)
                count = 0
                if len(names_days[f'day{names_x}_noon']) > 2:
                    for name in before_names["motsash"]:
                        if name in names_days[f'day{names_x}_noon'] and name not in max_out_names[0]:
                            if f'{name}0' in sequence_count.keys():
                                sequence_count[f'{name}0'] += 1
                                if sequence_count[f'{name}0'] >= max_seq0:
                                    max_out_names[0].append(name)
                            if count == 0:
                                weeks_dicts[num_week][f'Day{x}_1400'] = name
                            else:
                                weeks_dicts[num_week][f'Day{x}_1400'] += "\n" + name
                            names_days[f'day{names_x}_noon'].remove(name)
                            count += 1
                            if count == 2:
                                break
                    if count < 2:
                        for name in names_days[f'day{names_x}_noon']:
                            if count == 0:
                                weeks_dicts[num_week][f'Day{x}_1400'] = name
                            else:
                                weeks_dicts[num_week][f'Day{x}_1400'] += "\n" + name
                            names_days[f'day{names_x}_noon'].remove(name)
                            count += 1
                            if count == 2:
                                break
                else:
                    for name in before_names["motsash"]:
                        if name in names_days[f'day{names_x}_noon'] and name not in max_out_names[0]:
                            if f'{name}0' in sequence_count.keys():
                                sequence_count[f'{name}0'] += 1
                                if sequence_count[f'{name}0'] >= max_seq0:
                                    max_out_names[0].append(name)
                            weeks_dicts[num_week][f'Day{x}_1400'] = name
                            names_days[f'day{names_x}_noon'].remove(name)
                            count += 1
                            break
                    if count == 0:
                        for name in names_days[f'day{names_x}_noon']:
                            weeks_dicts[num_week][f'Day{x}_1400'] = name
                            names_days[f'day{names_x}_noon'].remove(name)
                            count += 1
                            break
                # noon
                insert_all_to_form(weeks_dicts[num_week], names_days[f'day{names_x}_noon'], x, "1500")
            # morning 630 and 700 search
            else:
                if x > 2:
                    temp_morning = []
                    for name in names_days[f'day{names_x}_morning']:
                        temp_morning.append(name)
                    chosen = search_and_put(weeks_dicts[num_week], names_days[f'day{names_x - 1}_noon'], names_days[f'day{names_x}_morning']
                                                 , x, "630", max_out_names[1], 1, sequence_count, max_seq0,
                                                 max_seq1, True, names_days[f'day{names_x - 2}_night'])
                else:
                    chosen = search_and_put(weeks_dicts[num_week], names_days[f'day{names_x - 1}_noon'],
                                                 names_days[f'day{names_x}_morning'], x,
                                                 "630", max_out_names[1], 1, sequence_count, max_seq0,
                                                 max_seq1, True, [])
                if not chosen:
                    temp_morning = seperate_list(names_days[f'day{names_x}_morning'], max_out_names)
                    if len(temp_morning) > 0:
                        chosen = insert_random(weeks_dicts[num_week], temp_morning, "630", x, 0)
                        if chosen is not None:
                            names_days[f'day{names_x}_morning'].remove(chosen)
                    else:
                        insert_random(weeks_dicts[num_week], names_days[f'day{names_x}_morning'], "630", x, 0)
                chosen = search_and_put(weeks_dicts[num_week], names_days[f'day{names_x - 1}_noon'], names_days[f'day{names_x}_morning'], x,
                                             "700_search", max_out_names[1], 1, sequence_count, max_seq0,
                                             max_seq1, False, [])
                if not chosen and len(names_days[f'day{names_x}_morning']) > 0:
                    chosen = insert_random(weeks_dicts[num_week], names_days[f'day{names_x}_morning'], "700_search", x, 0)
                # noon
                count = 0
                if len(names_days[f'day{names_x}_noon']) > 2:
                    for name in names_days[f'day{names_x}_noon']:
                        if name in names_days[f'day{names_x -1}_night'] and name not in max_out_names[0]:
                            if f'{name}0' in sequence_count.keys():
                                sequence_count[f'{name}0'] += 1
                                if sequence_count[f'{name}0'] >= max_seq0:
                                    max_out_names[0].append(name)
                            if count == 2:
                                break
                            if count == 0:
                                weeks_dicts[num_week][f'Day{x}_1400'] = name
                            else:
                                weeks_dicts[num_week][f'Day{x}_1400'] += "\n" + name
                            count += 1
                            names_days[f'day{names_x}_noon'].remove(name)
                    if count < 2:
                        while count < 2:
                            insert_random(weeks_dicts[num_week], names_days[f'day{names_x}_noon'], "1400", x, count)
                            count += 1
                else:
                    chosen = search_and_put(weeks_dicts[num_week], names_days[f'day{names_x -1}_night'],
                                                 names_days[f'day{names_x}_noon'], x,
                                                 "1400", max_out_names[0], 0, sequence_count, max_seq0,
                                                 max_seq1, True, [])
                    if not chosen:
                        insert_random(weeks_dicts[num_week], names_days[f'day{names_x}_noon'], "1400", x, 0)
                        chosen = True
                # noon
                insert_all_to_form(weeks_dicts[num_week], names_days[f'day{names_x}_noon'], x, "1500")
            # morning
            count = 0
            if not chosen:
                insert_random(weeks_dicts[num_week], names_days[f'day{names_x}_morning'], "700_search", x, 0)
            chosen = False
            for morning in range(len(names_days[f'day{names_x}_morning'])):
                r = random.randint(0, len(names_days[f'day{names_x}_morning']) - 1)
                if count < 3:
                    weeks_dicts[num_week][f'Day{x}_720_{count + 1}'] = names_days[f'day{names_x}_morning'][r]
                else:
                    weeks_dicts[num_week][f'Day{x}_720_3'] += "\n" + names_days[f'day{names_x}_morning'][r]
                names_days[f'day{names_x}_morning'].pop(r)
                count += 1
            # night
            insert_all_to_form(weeks_dicts[num_week], names_days[f'day{names_x}_night'], x, "2300")
        else:
            count = 0
            shift = "_700_search"
            # morning
            for name in names_days[f'day{names_x}_morning']:
                if count == 2:
                    shift = "_700_manager"
                if count == 0 or count == 2:
                    weeks_dicts[num_week][f'Day{x + 1}{shift}'] = name
                else:
                    weeks_dicts[num_week][f'Day{x + 1}{shift}'] += "\n" + name
                count += 1
            # noon
            insert_all_to_form(weeks_dicts[num_week], names_days[f'day{names_x}_noon'], x, "1500")
            # night
            insert_all_to_form(weeks_dicts[num_week], names_days[f'day{names_x}_night'], x, "2300")
        if days_count == 6:
            num_week -= 1
            days_count = -1
        days_count += 1
    return weeks_dicts


def search_and_put(weeks_dict, for_list, check_list, day, time, max_out_names, seq,
                   sequence_count, max_seq0, max_seq1, is_seq, extra_seq):
    if seq == 0:
        max_seq = max_seq0
    else:
        max_seq = max_seq1
    for name in for_list:
        if name in check_list and name not in extra_seq:
            if is_seq:
                if name not in max_out_names:
                    if f'{name}{seq}' in sequence_count.keys():
                        sequence_count[f'{name}{seq}'] += 1
                        if sequence_count[f'{name}{seq}'] >= max_seq:
                            max_out_names.append(name)
                    if weeks_dict.get(f'Day{day}_{time}', "") == "":
                        weeks_dict[f'Day{day}_{time}'] = name
                    else:
                        weeks_dict[f'Day{day}_{time}'] += "\n" + name
                    check_list.remove(name)
                    return True
            else:
                if weeks_dict.get(f'Day{day}_{time}', "") == "":
                    weeks_dict[f'Day{day}_{time}'] = name
                else:
                    weeks_dict[f'Day{day}_{time}'] += "\n" + name
                check_list.remove(name)
                return True
    return False


def insert_all_to_form(weeks_dict, for_list, day, time):
    weeks_dict[f'Day{day}_{time}'] = '\n'.join(for_list)
    for name in for_list:
        for_list.remove(name)


def seperate_list(shift, max_out_names):
    new_list = []
    for s in shift:
        if s not in max_out_names:
            new_list.append(s)
    return new_list


def insert_random(weeks_dict, list1, time, day, count):
    if len(list1) > 0:
        r = random.randint(0, len(list1) - 1)
        if count == 0:
            weeks_dict[f'Day{day}_{time}'] = list1[r]
        else:
            weeks_dict[f'Day{day}_{time}'] += "\n" + list1[r]
        return list1.pop(r)
    else:
        return None


def organization_valid(organization, request):
    organization1 = get_input(organization)
    input_days = {}
    valid = True
    keys = ["_630", "_700_manager", "_700_search", "_720_1", "_720_pull", "_720_2", "_720_3", "_1400", "_1500",
            "_1500_1900",
            "_2300"]
    for i in range(1, organization.num_weeks * 7 + 1):
        day = "day" + str(i)
        input_days[day + "M"] = []
        input_days[day + "A"] = []
        input_days[day + "N"] = []
        for x in range(len(keys)):
            if x < 7:
                input_days[day + "M"] += organization1[day + keys[x]].split("\n")
            elif x < 10:
                input_days[day + "A"] += organization1[day + keys[x]].split("\n")
            else:
                input_days[day + "N"] += organization1[day + keys[x]].split("\n")
    for key in input_days:
        for i in range(len(input_days[key])):
            input_days[key][i] = input_days[key][i].replace(" ", "")
            input_days[key][i] = input_days[key][i].replace("\n", "")
            input_days[key][i] = input_days[key][i].replace("\r", "")
    for key in input_days:
        for i in range(input_days[key].count('')):
            input_days[key].remove('')
    massages_sent = []
    for key in input_days:
        for name in input_days[key]:
            num_day = key.replace("day", "")
            num_day = num_day.replace("A", "")
            num_day = num_day.replace("N", "")
            num_day = num_day.replace("M", "")
            message1 = name + " ביום ה-" + num_day + " בשתי משמרות רצופות"
            message2 = name + " ביום ה-" + num_day + " באותה משמרת פעמיים"
            day = "day" + num_day
            day_before = "day" + str(int(num_day) - 1)
            day_after = "day" + str(int(num_day) + 1)
            if name in input_days[day + "M"]:
                if is_more_than_once(input_days[day + "M"], name):
                    if message2 not in massages_sent:
                        messages.info(request, translate_text(message2, request.user, "hebrew"))
                        massages_sent.append(message2)
                    valid = False
                if int(num_day) != 1:
                    if name in input_days[day + "A"] or name in input_days[day_before + "N"]:
                        if message1 not in massages_sent:
                            messages.info(request, translate_text(message1, request.user, "hebrew"))
                            massages_sent.append(message1)
                        valid = False
                else:
                    if name in input_days[day + "A"]:
                        if message1 not in massages_sent:
                            messages.info(request, translate_text(message1, request.user, "hebrew"))
                            massages_sent.append(message1)
                        valid = False
            if name in input_days[day + "A"]:
                if is_more_than_once(input_days[day + "A"], name):
                    if message2 not in massages_sent:
                        messages.info(request, translate_text(message2, request.user, "hebrew"))
                        massages_sent.append(message2)
                    valid = False
                if name in input_days[day + "M"] or name in input_days[day + "N"]:
                    if message1 not in massages_sent:
                        messages.info(request, translate_text(message1, request.user, "hebrew"))
                        massages_sent.append(message1)
                    valid = False
            if name in input_days[day + "N"]:
                if is_more_than_once(input_days[day + "N"], name):
                    if message2 not in massages_sent:
                        messages.info(request, translate_text(message2, request.user, "hebrew"))
                        massages_sent.append(message2)
                    valid = False
                if int(num_day) != organization.num_weeks * 7:
                    if name in input_days[day + "A"] or name in input_days[day_after + "M"]:
                        if message1 not in massages_sent:
                            messages.info(request, translate_text(message1, request.user, "hebrew"))
                            massages_sent.append(message1)
                        valid = False
                else:
                    if name in input_days[day + "A"]:
                        if message1 not in massages_sent:
                            messages.info(request, translate_text(message1, request.user, "hebrew"))
                            massages_sent.append(message1)
                        valid = False
    if valid:
        messages.success(request, translate_text("סידור תקין", request.user, "hebrew"))

def to_week_form(form, request, j):
    for i in range(1, 8):
        setattr(form.instance, f"Day{i}_630", request.POST.get(f"day{i}_630_{j}"))
        setattr(form.instance, f"Day{i}_700_search", request.POST.get(f"day{i}_700_search_{j}"))
        setattr(form.instance, f"Day{i}_700_manager", request.POST.get(f"day{i}_700_manager_{j}"))
        setattr(form.instance, f"Day{i}_720_1", request.POST.get(f"day{i}_720_1_{j}"))
        setattr(form.instance, f"Day{i}_720_pull", request.POST.get(f"day{i}_720_pull_{j}"))
        setattr(form.instance, f"Day{i}_720_2", request.POST.get(f"day{i}_720_2_{j}"))
        setattr(form.instance, f"Day{i}_720_3", request.POST.get(f"day{i}_720_3_{j}"))
        setattr(form.instance, f"Day{i}_1400", request.POST.get(f"day{i}_1400_{j}"))
        setattr(form.instance, f"Day{i}_1500_1900", request.POST.get(f"day{i}_1500_1900_{j}"))
        setattr(form.instance, f"Day{i}_1500", request.POST.get(f"day{i}_1500_{j}"))
        setattr(form.instance, f"Day{i}_2300", request.POST.get(f"day{i}_2300_{j}"))
        setattr(form.instance, f"Day{i}_notes", request.POST.get(f"day{i}_notes_{j}"))


def is_more_than_once(list, name):
    num = 0
    for n in list:
        if n == name:
            num += 1
    if num > 1:
        return True
    return False


def check_if_in_list(names, name):
    if name in names:
        return True
    return None


class OrganizationListView(LoginRequiredMixin, ListView):
    model = Organization
    template_name = "Schedule/organizations_list.html"
    context_object_name = "organizations"
    ordering = ["-date"]
    paginate_by = 1

    def get_context_data(self, **kwargs):
        ctx = super(OrganizationListView, self).get_context_data(**kwargs)
        weeks = Week.objects.all()
        ctx["weeks"] = weeks
        return ctx


def WriteToExcel(served, notes, notes_general, dates, user):
    # Create a workbook and add a worksheet.
    buffer = io.BytesIO()
    workbook = xlsxwriter.Workbook(buffer)
    worksheet = workbook.add_worksheet()
    worksheet.right_to_left()

    days = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"]
    user_settings = USettings.objects.all().filter(user=user).first()
    if user_settings.language == 'english':
        days = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

    maxes = {"morning": 0, "after_noon": 0, "night": 0}

    for key in served:
        temp = 0
        if key.count("M"):
            split = served[key].split("\n")
            for x in range(len(split)):
                if split[x] != "(לא משיכה)":
                    temp += 1
            if temp > maxes["morning"]:
                maxes["morning"] = temp
        elif key.count("A"):
            split = served[key].split("\n")
            if len(split) > maxes["after_noon"]:
                maxes["after_noon"] = len(split)
        else:
            split = served[key].split("\n")
            if len(split) > maxes["night"]:
                maxes["night"] = len(split)

    maxes["morning"] += 1
    maxes["after_noon"] += 2
    maxes["night"] += 2

    # Write a total using a formula.
    title_format = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'font_size': 24,
        'fg_color': 'white'})
    cell_format = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
    })
    cell_no_pull_format = workbook.add_format({
        'font_color': "#ff0000"
    })
    border_bottom_format = workbook.add_format({
        'bottom': 5,
        'bottom_color': '#000000'
    })
    border_left_format = workbook.add_format({
        'left': 5,
        'left_color': '#000000'
    })
    border_left_bottom_format = workbook.add_format({
        'left': 5,
        'left_color': '#000000',
        'bottom': 5,
        'bottom_color': '#000000',
    })
    border_right_format = workbook.add_format({
        'right': 5,
        'right_color': '#000000'
    })
    border_right_bottom_format = workbook.add_format({
        'right': 5,
        'right_color': '#000000',
        'bottom': 5,
        'bottom_color': '#000000',
    })
    # Building first Structure
    col = 0
    for x in range(len(dates) + 1):
        worksheet.write(4 + maxes["morning"], col, None, border_bottom_format)
        worksheet.write(4 + maxes["after_noon"] + maxes["morning"], col, None, border_bottom_format)
        col += 1
    row = 0
    sum_maxes = maxes["morning"] + maxes["after_noon"] + maxes["night"] + 6
    for i in range(int(len(dates) / 7)):
        row = 0
        for x in range(sum_maxes):
            if x == 4 + maxes["morning"] or x == 4 + maxes["morning"] + maxes["after_noon"]:
                worksheet.write(row, 7 + 7 * i, None, border_right_bottom_format)
            else:
                worksheet.write(row, 7 + 7 * i, None, border_right_format)
            row += 1
    # Building second Structure
    worksheet.merge_range('A1:H2', translate_text('הגשות', user, "hebrew"), title_format)
    worksheet.merge_range('I1:P2', translate_text('הגשות', user, "hebrew"), title_format)
    worksheet.merge_range('Q1:X2', dates[0].strftime("%d.%m") + "-" + dates[-1].strftime("%d.%m"),
                          title_format)
    worksheet.write(2, 0, translate_text("תאריך", user, "hebrew"), cell_format)
    col = 1
    for d in dates:
        worksheet.write(2, col, d.strftime("%d.%m"), cell_format)
        col += 1
    worksheet.write(3, 0, translate_text("יום", user, "hebrew"), cell_format)
    col = 1
    for i in range(int(len(dates) / 7)):
        for d in days:
            worksheet.write(3, col, d, cell_format)
            col += 1
    worksheet.merge_range(f'A5:A{5 + maxes["morning"]}', translate_text('בוקר', user, "hebrew"), cell_format)
    worksheet.merge_range(
        f'A{5 + maxes["morning"] + 1}:A{5 + maxes["morning"] + maxes["after_noon"]}', translate_text('צהריים', user, "hebrew"), cell_format)
    worksheet.merge_range(
        f'A{5 + maxes["morning"] + maxes["after_noon"] + 1}:A{5 + maxes["morning"] + maxes["after_noon"] + maxes["night"]}',
        translate_text('לילה', user, "hebrew"), cell_format)
    start_extra = len(dates) + 3
    worksheet.merge_range(f'{get_column_letter(start_extra)}4:{get_column_letter(start_extra + 1)}4', 'שם', cell_format)
    worksheet.write(f"{get_column_letter(start_extra + 2)}4", translate_text('בוקר', user, "hebrew") + " 1", cell_format)
    worksheet.write(f"{get_column_letter(start_extra + 3)}4", translate_text('בוקר', user, "hebrew") + " 2", cell_format)
    worksheet.write(f"{get_column_letter(start_extra + 4)}4", translate_text('צהריים', user, "hebrew") + " 1", cell_format)
    worksheet.write(f"{get_column_letter(start_extra + 5)}4", translate_text('צהריים', user, "hebrew") + " 2", cell_format)
    worksheet.write(f"{get_column_letter(start_extra + 6)}4", translate_text('לילה', user, "hebrew"), cell_format)
    worksheet.write(f"{get_column_letter(start_extra + 7)}4", translate_text("סופ\"ש", user, "hebrew"), cell_format)

    # Adding Data
    users = []
    row = 4
    col = 1
    for key in served:
        print(key)
        if key.count("M"):
            row = 4
            split = served[key].split("\n")
            for x in range(len(split)):
                if split[x] != "(לא משיכה)":
                    if x + 1 < len(split):
                        if split[x + 1] == "(לא משיכה)":
                            worksheet.write(row, col, split[x], cell_no_pull_format)
                        else:
                            worksheet.write(row, col, split[x])
                    else:
                        worksheet.write(row, col, split[x])
                    if split[x] not in users:
                        users.append(split[x])
                    row += 1
        elif key.count("A"):
            row = 4 + maxes["morning"] + 1
            split = served[key].split("\n")
            for x in range(len(split)):
                worksheet.write(row, col, split[x])
                if split[x] not in users:
                    users.append(split[x])
                row += 1
        else:
            row = 4 + maxes["morning"] + maxes["after_noon"] + 1
            split = served[key].split("\n")
            for x in range(len(split)):
                worksheet.write(row, col, split[x])
                if split[x] not in users:
                    users.append(split[x])
                row += 1
            col += 1

    num_rows = len(users) + 1
    for x in range(num_rows):
        col = len(dates) + 4
        if x == 0:
            worksheet.merge_range(f'{get_column_letter(col - 1)}{4 + x + 1}:{get_column_letter(col)}{4 + x + 1}', '', cell_format)
        else:
            worksheet.merge_range(f'{get_column_letter(col - 1)}{4 + x + 1}:{get_column_letter(col)}{4 + x + 1}', users[x - 1], cell_format)
        for c in range(6):
            worksheet.write(4 + x, col + c, "", cell_format)
    worksheet.merge_range(f'{get_column_letter(start_extra)}{4 + num_rows + 1}:{get_column_letter(start_extra + 1)}{4 + num_rows + 1}',
                          translate_text('סה\"כ', user, "hebrew"), cell_format)
    worksheet.write(f'{get_column_letter(start_extra + 2)}{4 + num_rows + 1}', f'=SUM({get_column_letter(start_extra + 2)}5:{get_column_letter(start_extra + 2)}{4 + num_rows})', cell_format)
    worksheet.write(f'{get_column_letter(start_extra + 3)}{4 + num_rows + 1}', f'=SUM({get_column_letter(start_extra + 3)}5:{get_column_letter(start_extra + 3)}{4 + num_rows})', cell_format)
    worksheet.write(f'{get_column_letter(start_extra + 4)}{4 + num_rows + 1}', f'=SUM({get_column_letter(start_extra + 4)}5:{get_column_letter(start_extra + 4)}{4 + num_rows})', cell_format)
    worksheet.write(f'{get_column_letter(start_extra + 5)}{4 + num_rows + 1}', f'=SUM({get_column_letter(start_extra + 5)}5:{get_column_letter(start_extra + 5)}{4 + num_rows})', cell_format)
    worksheet.write(f'{get_column_letter(start_extra + 6)}{4 + num_rows + 1}', f'=SUM({get_column_letter(start_extra + 6)}5:{get_column_letter(start_extra + 6)}{4 + num_rows})', cell_format)
    worksheet.write(f'{get_column_letter(start_extra + 7)}{4 + num_rows + 1}', f'=SUM({get_column_letter(start_extra + 7)}5:{get_column_letter(start_extra + 7)}{4 + num_rows})', cell_format)

    row = 4 + num_rows + 4

    worksheet.merge_range(f'{get_column_letter(start_extra)}{row}:{get_column_letter(start_extra + 7)}{row + 3}', translate_text('משמרות לאיכות', user, "hebrew"), title_format)
    worksheet.merge_range(f'{get_column_letter(start_extra)}{row + 4}:{get_column_letter(start_extra + 7)}{row + 5}', translate_text('שבוע ראשון', user, "hebrew"), title_format)
    worksheet.merge_range(f'{get_column_letter(start_extra)}{row + 6}:{get_column_letter(start_extra + 7)}{row + 6}', '', cell_format)
    worksheet.merge_range(f'{get_column_letter(start_extra)}{row + 7}:{get_column_letter(start_extra + 7)}{row + 7}', '', cell_format)
    worksheet.merge_range(f'{get_column_letter(start_extra)}{row + 8}:{get_column_letter(start_extra + 7)}{row + 9}', translate_text('שבוע שני', user, "hebrew"), title_format)
    worksheet.merge_range(f'{get_column_letter(start_extra)}{row + 10}:{get_column_letter(start_extra + 7)}{row + 10}', '', cell_format)
    worksheet.merge_range(f'{get_column_letter(start_extra)}{row + 11}:{get_column_letter(start_extra + 7)}{row + 11}', '', cell_format)

    row = row + 13
    count = 0
    worksheet.merge_range(
        f'{get_column_letter(start_extra)}{row + count}:{get_column_letter(start_extra + 7)}{row + count + 1}',
        translate_text('הערות', user, "hebrew"), title_format)
    count += 2
    split = notes_general.split("\n")
    if len(split) > 0:
        for s in split:
            worksheet.merge_range(
                f'{get_column_letter(start_extra)}{row + count}:{get_column_letter(start_extra + 7)}{row + count}', s,
                cell_format)
            count += 1
    for n in range(len(notes)):
        worksheet.merge_range(f'{get_column_letter(start_extra)}{row + count}:{get_column_letter(start_extra + 7)}{row + count}', translate_text(f'שבוע {str(n + 1)}', user, "hebrew"), title_format)
        count += 1
        split = notes[n].split("\n")
        if len(split) > 0:
            for s in split:
                worksheet.merge_range(f'{get_column_letter(start_extra)}{row + count}:{get_column_letter(start_extra + 7)}{row + count}', s, cell_format)
                count += 1

    worksheet.merge_range(f'{get_column_letter(start_extra + 10)}4:{get_column_letter(start_extra + 15)}5', translate_text('אירועים', user, "hebrew"), title_format)
    events = Event.objects.all()
    events_notes = []
    temp = ""
    for x in range(len(dates)):
        if len(events.filter(date2=dates[x])) > 0:
            for ev in events.filter(date2=dates[x]):
                if ev.nickname != "כולם":
                    events_notes.append(translate_text(f'בתאריך {ev.date2} יש {ev.description} ל{ev.nickname}', user, "hebrew"))
                else:
                    events_notes.append(translate_text(f'בתאריך {ev.date2} יש {ev.description}', user, "hebrew"))
    row = 6
    count = 0
    for s in events_notes:
        worksheet.merge_range(f'{get_column_letter(start_extra + 10)}{row + count}:{get_column_letter(start_extra + 15)}{row + count}', s, cell_format)
        count += 1

    workbook.close()
    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    file_name = "serve" + dates[0].strftime("%d.%m")
    return FileResponse(buffer, as_attachment=True, filename=f'{file_name}.xlsx')


def number_to_day2(num):
    day = "יום "
    if num == 1:
        return day + "ראשון"
    elif num == 2:
        return day + "שני"
    elif num == 3:
        return day + "שלישי"
    elif num == 4:
        return day + "רביעי"
    elif num == 5:
        return day + "חמישי"
    elif num == 6:
        return day + "שישי"
    else:
        return day + "שבת"


guards_num = {}
for x in range(14):
    guards_num[f"M{x}"] = 5
    guards_num[f"A{x}"] = 3
    guards_num[f"N{x}"] = 1
guards_num["A13"] = 0
guards_num["A12"] = 0
guards_num["M12"] = 0
guards_num["A6"] = 0
guards_num["M5"] = 0
guards_num["A5"] = 0
guards_num["N13"] = 2
guards_num["N12"] = 2
guards_num["M13"] = 2
guards_num["N6"] = 2
guards_num["N5"] = 2
guards_num["M6"] = 2


def compare_organizations(served, guards_num, organization, officer, sat_night, users, users_settings):
    organizer = Organizer(served, guards_num, organization, officer, sat_night, users, users_settings)
    organizer.organize()
    for i in range(1):
        new_organizer = Organizer(served, guards_num, organization, officer, sat_night, users, users_settings)
        new_organizer.organize()
        if organizer.notes > new_organizer.notes:
            organizer = new_organizer
        if organizer.notes == 0:
            break
    return organizer


class OrganizationSuggestionView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Organization
    template_name = "Schedule/Suggestion.html"

    def get_context_data(self, **kwargs):
        ctx = super(OrganizationSuggestionView, self).get_context_data(**kwargs)
        settings = Settings.objects.all().first()
        try:
            last_organization = Organization.objects.all().order_by('-date')[get_num_organization(self.get_object()) - 1]
        except:
            last_organization = ""
        users = User.objects.all()
        users_settings = USettings.objects.all()
        days = []
        for x in range(self.get_object().num_weeks * 7):
            days.append(self.get_object().date + datetime.timedelta(days=x))
        ctx["days"] = days
        served, notes = self.get_served()
        if last_organization != "":
            organizer = compare_organizations(served, guards_num, self.get_object(), settings.officer,
                                          last_organization.Day14_2300.split("\n"), users, users_settings)
        else:
            organizer = compare_organizations(served, guards_num, self.get_object(), settings.officer,
                                              "", users, users_settings)
        organized_str = {}
        for key in organizer.organized:
            organized_str[key] = '\n'.join(organizer.organized[key])
        ctx["organized"] = organized_str
        ctx["notes"] = organizer.notes
        ctx["guardsnumbers"] = guards_num
        return ctx

    def post(self, request, *args, **kwargs):
        for x in range(14):
            day = f'M{x}'
            if x == 6 or x == 13:
                guards_num[day] = int(self.request.POST.get(day, 2))
            elif x == 5 or x == 12:
                guards_num[day] = int(self.request.POST.get(day, 0))
            else:
                guards_num[day] = int(self.request.POST.get(day, 5))
            day = f'A{x}'
            if x == 5 or x == 6 or x == 12 or x == 13:
                guards_num[day] = int(self.request.POST.get(day, 0))
            else:
                guards_num[day] = int(self.request.POST.get(day, 3))
            day = f'N{x}'
            if x == 5 or x == 6 or x == 12 or x == 13:
                guards_num[day] = int(self.request.POST.get(day, 2))
            else:
                guards_num[day] = int(self.request.POST.get(day, 1))
        settings = Settings.objects.all().first()
        try:
            last_organization = Organization.objects.all().order_by('-date')[get_num_organization(self.get_object()) - 1]
        except:
            last_organization = ""
        users = User.objects.all()
        users_settings = USettings.objects.all()
        days = []
        for x in range(self.get_object().num_weeks * 7):
            days.append(self.get_object().date + datetime.timedelta(days=x))
        served, notes = self.get_served()
        if last_organization != "":
            organizer = compare_organizations(served, guards_num, self.get_object(), settings.officer,
                                              last_organization.Day14_2300.split("\n"), users, users_settings)
        else:
            organizer = compare_organizations(served, guards_num, self.get_object(), settings.officer,
                                              "", users, users_settings)
        if 'organize' in request.POST:
            return HttpResponseRedirect(self.request.path_info)
        elif 'excel' in request.POST:
            return organizer.WriteToExcel(notes, days, self.request.user)

    def get_served(self):
        served = {}
        for i in range(14):
            served["M" + str(i)] = []
            served["A" + str(i)] = []
            served["N" + str(i)] = []
        shifts_served = Shift.objects.all().filter(date=self.get_object().date)
        notes = {"general": "", "week1": "", "week2": ""}
        for shift in shifts_served:
            user = User.objects.all().filter(username=shift.username).first()
            user_settings = USettings.objects.all().filter(user=user).first()
            name = user_settings.nickname
            shifts = [shift.M1, shift.A1, shift.N1, shift.M2, shift.A2, shift.N2, shift.M3,
                      shift.A3, shift.N3, shift.M4, shift.A4, shift.N4, shift.M5, shift.A5, shift.N5,
                      shift.M6, shift.A6, shift.N6, shift.M7, shift.A7, shift.N7, shift.M8,
                      shift.A8, shift.N8, shift.M9, shift.A9, shift.N9, shift.M10, shift.A10,
                      shift.N10, shift.M11, shift.A11, shift.N11, shift.M12, shift.A12, shift.N12,
                      shift.M13, shift.A13, shift.N13, shift.M14, shift.A14, shift.N14]
            kind = "M"
            count = 0
            index = 0
            for s in shifts:
                if s:
                    served[kind + str(index)].append(user_settings.nickname)
                count = count + 1
                if count == 0:
                    kind = "M"
                elif count == 1:
                    kind = "A"
                elif count == 2:
                    kind = "N"
                else:
                    kind = "M"
                    index = index + 1
                    count = 0
            notes1 = [shift.notes1, shift.notes2, shift.notes3,
                      shift.notes4, shift.notes5, shift.notes6, shift.notes7]
            index = 1
            for n in notes1:
                if n != "":
                    notes["week1"] = notes["week1"] + name + ": " \
                                     + number_to_day2(index) + " - " + n + "\n"
                index += 1
            notes2 = [shift.notes8, shift.notes9, shift.notes10,
                      shift.notes11, shift.notes12, shift.notes13, shift.notes14]
            index = 1
            for n in notes2:
                if n != "":
                    notes["week2"] = notes["week2"] + name + ": " \
                                     + number_to_day2(index) + " - " + n + "\n"
                index += 1
            if shift.notes != "":
                notes["general"] = notes["general"] + name + ": " + shift.notes + "\n"
        return served, notes

    def test_func(self):
        if self.request.user.is_staff:
            return True
        return False


# filters

@register.filter
def getmonth(month):
    letter = month.lower()[0]
    if (letter >= 'a' and letter <= 'z'):
        return month
    else:
        months_he = ["ינו", "פבר", "מרץ", "אפר", "מאי", "יונ", "יול", "אוג", "ספט", "אוק", "נוב", "דצמ"]
        months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
        return months[months_he.index(month)]


@register.filter
def getfullname(user):
    return user.first_name + " " + user.last_name

@register.filter
def getday(string):
    letter = datetime.datetime.now().strftime("%b").lower()[0]
    if (letter >= 'a' and letter <= 'z'):
        return datetime.datetime.now()
    return datetime.datetime.now()

@register.filter
def edit_permission(user, name):
    username = user.first_name + " " + user.last_name
    if user.groups.filter(name="manager").exists() or username == name:
        return True
    return False

@register.filter
def timestr(time):
    if type(time) == Time:
        return time.strftime("%H:%M")
    else:
        return time

@register.filter
def num_to_shift(num):
    if num == 1:
        return "בוקר"
    elif num == 2:
        return "צהריים"
    return "לילה"

@register.filter
def counter_shifts(counter, arming_logs):
    morning = 0
    afternoon = 0
    night = 0
    for log in arming_logs:
        if log.shift_num == 1:
            morning += 1
        elif log.shift_num == 2:
            afternoon += 1
        else:
            night += 1
    if counter <= morning:
        return counter
    elif counter <= morning + afternoon:
        return counter - morning
    else:
        return counter - morning - afternoon

@register.filter(name="translate_text")
def translate_text(text, user, from_language="hebrew"):
    if user.is_authenticated:
        user_settings = USettings.objects.all().filter(user=user).first()
        if from_language != user_settings.language:
            langs_dict = GoogleTranslator.get_supported_languages(as_dict=True)
            translator = GoogleTranslator(source='auto', target=langs_dict[user_settings.language])
            return translator.translate(text).capitalize()
    return text


@register.filter
def translate_text_batch(texts, user, from_language):
    translated = []
    if user.is_authenticated:
        user_settings = USettings.objects.all().filter(user=user).first()
        if from_language != user_settings.language:
            langs_dict = GoogleTranslator.get_supported_languages(as_dict=True)
            translator = GoogleTranslator(source='auto', target=langs_dict[user_settings.language])
            for text in texts:
                translated.append(translator.translate(text).capitalize())
            return translated
    return texts


@register.filter
def get_base_string(user, num):
    return translate_text(base_strings[int(num)], user, "hebrew")


@register.filter
def get_user_name(username):
    user = User.objects.all().filter(username=username).first()
    name = user.first_name + " " + user.last_name
    return name


@register.filter
def get_item(dictionary, key):
    return dictionary.get(key)


@register.filter
def is_divided5(num):
    if num % 5 == 0:
        return True
    return False


@register.filter
def get_index(list, item):
    return list.index(item)


@register.filter
def minus(item, num):
    return item - num


@register.filter
def plus_days(date):
    return date + datetime.timedelta(days=13)


@register.filter
def cut_list(list, num):
    new_list = []
    if num == 1:
        for x in range(7):
            new_list.append(list[x])
    else:
        for x in range(7, len(list)):
            new_list.append(list[x])
    return new_list


@register.filter
def is_in_td(text, nickname):
    if text is not None and len(text) > 0:
        split = text.split("\n")
        for s in split:
            s = s.replace(" ", "")
            s = s.replace("\n", "")
            s = s.replace("\r", "")
            if nickname == s:
                return True
    return False


@register.filter
def is_string(item):
    return isinstance(item, str)


@register.filter
def clip_dictionary(served, num_week):
    days = range((num_week * 7) + 1, (num_week * 7) + 8)
    new_list = []
    for day in days:
        new_list.append(served[f'day{day}'])
    return new_list


@register.filter
def clip_dictionary_served1(served, kind):
    new_list = {}
    for key in served.keys():
        if kind in key:
            new_list[key] = served[key]
    return new_list


@register.filter
def clip_dictionary_served2(served, num_week):
    kind = "M"
    for key in served.keys():
        if kind in key:
            break
        elif 'A' in key:
            kind = 'A'
            break
        elif 'N' in key:
            kind = 'N'
            break
    days = range((num_week * 7) + 1, (num_week * 7) + 8)
    new_list = []
    for day in days:
        new_list.append(served[kind + str(day)])
    return new_list


@register.filter
def clip_days(days, num_week):
    new_days = []
    for i in range(num_week * 7, num_week * 7 + 7):
        new_days.append(days[i])
    return new_days


@register.filter
def to_array(end, start):
    start = int(start)
    end = int(end)
    return range(start, end)


@register.filter
def to_array2(start, end):
    start = int(start)
    end = int(end)
    return range(start, end)


@register.filter
def get_form_data(form, kind):
    array = []
    for i in range(1, 8):
        array.append(getattr(form.instance, f"{kind}{i}"))
    return array


@register.filter
def get_days(organization):
    days = []
    for x in range(organization.num_weeks * 7):
        days.append(organization.date + datetime.timedelta(days=x))
    return days


@register.filter
def get_weeks(organization):
    weeks = Week.objects.all().filter(date=organization.date)
    return weeks


@register.filter
def get_num_organization(organization):
    organizations = Organization.objects.all().order_by('-date')
    i = 0
    for org in organizations:
        if org.date == organization.date:
            return i
        i += 1
    return -1
